"""ECM Simultaneous Heating/Cooling Savings Calculator — generates .xlsx with working Excel formulas.

Covers dual-fuel waste from simultaneous heating and cooling:
  - Heating season gas waste (HW fighting CHW)
  - Cooling season electrical waste (chiller removing heat from HW)
  - Cooling season gas waste (HW during cooling)
  - Valve characteristic corrections (equal-%, linear)

Usage:
  # CLI args (quick, single unit)
  py -3.12 ecm-simhtgclg-savings.py --name "RTU-2" --htg-mbh 592 --clg-tons 24 --hw-valve-pos 0.42 --sim-fraction 0.482 --elec-rate 0.10 --gas-rate 0.90 --output simhtgclg.xlsx

  # JSON config (reproducible, full detail)
  py -3.12 ecm-simhtgclg-savings.py --config inputs.json --output simhtgclg.xlsx

JSON config format:
{
  "project": "Example Project",
  "ecm_id": "ECM-11",
  "ecm_title": "RTU-2 Simultaneous Heating/Cooling Elimination",
  "elec_rate": 0.10,
  "gas_rate": 0.90,
  "chiller_eer": 10.5,
  "hours_per_week": 168,
  "weeks_per_year": 52,
  "cooling_season_weeks": 22,
  "unit_name": "RTU-2",
  "heating_capacity_mbh": 592,
  "cooling_capacity_tons": 24,
  "hw_valve_avg_position": 0.419,
  "chw_valve_avg_position": 0.968,
  "simultaneous_fraction": 0.482,
  "heating_waste_fraction": 0.10,
  "heating_waste_note": "5-15% est; using 10% midpoint",
  "valve_scenarios": [
    {"name": "Aggressive (no correction)", "assumed_flow_pct": 0.42, "note": "Valve position = flow (linear)"},
    {"name": "Adjusted (equal-% valve)", "assumed_flow_pct": 0.15, "note": "42% position ~ 15% flow"},
    {"name": "Conservative (equal-% + seasonal)", "assumed_flow_pct": 0.10, "note": "10% flow + seasonal adj"}
  ],
  "trend_data": {
    "period": "Jan 31 - Feb 1, 2026",
    "records": 2018,
    "source": "Phase 3.5 trend data"
  },
  "verifications": ["RTU-2 CHW valve position during cooling season"],
  "implementation_cost": "CHW valve actuator replacement (~$500-$1,500)"
}
"""
import argparse, json, sys
from pathlib import Path
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent))
from _styles import *


def build_workbook(cfg):
    wb = Workbook()
    elec_rate = cfg.get("elec_rate", 0.10)
    gas_rate = cfg.get("gas_rate", 0.90)
    eer = cfg.get("chiller_eer", 10.5)
    hr_wk = cfg.get("hours_per_week", 168)
    wk_yr = cfg.get("weeks_per_year", 52)
    cool_wk = cfg.get("cooling_season_weeks", 22)
    unit_name = cfg.get("unit_name", "Unit")
    htg_mbh = cfg.get("heating_capacity_mbh", 592)
    clg_tons = cfg.get("cooling_capacity_tons", 24)
    hw_valve = cfg.get("hw_valve_avg_position", 0.42)
    chw_valve = cfg.get("chw_valve_avg_position", 0.97)
    sim_frac = cfg.get("simultaneous_fraction", 0.48)
    htg_waste_frac = cfg.get("heating_waste_fraction", 0.10)
    htg_waste_note = cfg.get("heating_waste_note", "5-15% est; using 10% midpoint")
    scenarios = cfg.get("valve_scenarios", [
        {"name": "Aggressive (no correction)", "assumed_flow_pct": hw_valve, "note": "Valve position = flow"},
        {"name": "Adjusted (equal-% valve)", "assumed_flow_pct": 0.15, "note": "Equal-% characteristic"},
        {"name": "Conservative (equal-% + seasonal)", "assumed_flow_pct": 0.10, "note": "Adjusted + seasonal"}
    ])
    n_scen = len(scenarios)
    ecm_id = cfg.get("ecm_id", "ECM")
    ecm_title = cfg.get("ecm_title", "Simultaneous Heating/Cooling")
    project = cfg.get("project", "")
    trend = cfg.get("trend_data", {})

    # ── ASSUMPTIONS SHEET ──
    ws = wb.active
    ws.title = "Assumptions"
    set_col_widths(ws, [40, 18, 14, 50])
    section_header(ws, 1, "Assumptions & Constants", 4)
    for i, h in enumerate(["Parameter", "Value", "Units", "Basis / Source"], 1):
        ws.cell(row=2, column=i, value=h)
    style_header_row(ws, 2, 4)

    a_rows = [
        (3, "Electricity rate", elec_rate, "$/kWh", cfg.get("elec_rate_source", "Verify against utility bills"), dollar_fmt2),
        (4, "Natural gas rate", gas_rate, "$/therm", cfg.get("gas_rate_source", "Verify against utility bills"), dollar_fmt2),
        (5, "Chiller EER", eer, "", cfg.get("eer_source", "Chiller nameplate"), dec1_fmt),
        (6, "Hours per week", hr_wk, "hr/wk", "24 hr/day x 7 days/wk", num_fmt),
        (7, "Weeks per year", wk_yr, "wk/yr", "", num_fmt),
        (8, "Cooling season weeks", cool_wk, "weeks", cfg.get("cool_season_note", "~May through September"), num_fmt),
    ]
    for row, param, val, units, basis, fmt in a_rows:
        ws.cell(row=row, column=1, value=param)
        input_cell(ws, row, 2, val, fmt)
        ws.cell(row=row, column=3, value=units)
        ws.cell(row=row, column=4, value=basis).font = note_font
    # Formulas
    r = 9
    ws.cell(row=r, column=1, value="Non-cooling season weeks")
    formula_cell(ws, r, 2, "=B7-B8", num_fmt)
    ws.cell(row=r, column=3, value="weeks")
    style_range(ws, 3, 9, 4)
    add_color_legend(ws, 11)

    # ── CALCULATIONS SHEET ──
    ws_c = wb.create_sheet("Calculations")
    total_cols = 2 + n_scen  # label + n scenarios + units
    set_col_widths(ws_c, [42] + [18] * n_scen + [14, 40])
    section_header(ws_c, 1, f"{ecm_id}: {ecm_title}", total_cols + 1)

    # Equipment data
    r = 3
    ws_c.cell(row=r, column=1, value=f"{unit_name} Equipment Data").font = subsection_font
    if trend:
        ws_c.cell(row=r, column=2, value=f"({trend.get('source', '')}, {trend.get('period', '')})").font = note_font

    r = 4
    for i, h in enumerate(["Parameter", "Value", "Units", "Source"], 1):
        ws_c.cell(row=r, column=i, value=h)
    style_header_row(ws_c, r, 4)

    equip_data = [
        (5, "Heating capacity", htg_mbh, "MBH", ""),
        (6, "Cooling capacity", clg_tons, "tons", ""),
        (7, "HW valve avg position", hw_valve, "", f"HW valve position from trend data"),
        (8, "CHW valve avg position", chw_valve, "", f"CHW valve position from trend data"),
        (9, "Simultaneous operation fraction", sim_frac, "", f"{sim_frac*100:.1f}% of {trend.get('records', 'N')} records"),
    ]
    for row, param, val, units, src in equip_data:
        ws_c.cell(row=row, column=1, value=param)
        fmt = pct_fmt if isinstance(val, float) and val < 1 else (num_fmt if isinstance(val, int) else dec1_fmt)
        input_cell(ws_c, row, 2, val, fmt)
        ws_c.cell(row=row, column=3, value=units)
        ws_c.cell(row=row, column=4, value=src).font = note_font
    style_range(ws_c, 5, 9, 4)

    # Valve correction scenarios
    r = 11
    ws_c.cell(row=r, column=1, value="Valve Characteristic Assumptions").font = subsection_font
    ws_c.cell(row=r + 1, column=1, value="Equal-percentage valve: position != flow. Correction factors below:").font = note_font

    r = 13
    for i, h in enumerate(["Scenario", "Valve Position", "Assumed Flow %", "Correction Factor", "Basis"], 1):
        ws_c.cell(row=r, column=i, value=h)
    style_header_row(ws_c, r, 5)

    scen_start = 14
    for s_idx, scen in enumerate(scenarios):
        r = scen_start + s_idx
        ws_c.cell(row=r, column=1, value=scen["name"])
        formula_cell(ws_c, r, 2, "=B7", pct_fmt)
        input_cell(ws_c, r, 3, scen["assumed_flow_pct"], pct_fmt)
        formula_cell(ws_c, r, 4, f"=C{r}/B7", dec2_fmt)
        ws_c.cell(row=r, column=5, value=scen.get("note", "")).font = note_font
    style_range(ws_c, scen_start, scen_start + n_scen - 1, 5)

    # ── Heating Season Gas Waste ──
    r = scen_start + n_scen + 1
    ws_c.cell(row=r, column=1, value="Heating Season Gas Waste").font = subsection_font
    htg_section_start = r + 1
    r = htg_section_start
    htg_headers = ["Calculation Step"] + [s["name"] for s in scenarios] + ["Units"]
    for i, h in enumerate(htg_headers, 1):
        ws_c.cell(row=r, column=i, value=h)
    style_header_row(ws_c, r, len(htg_headers))

    # Est HW output
    r = htg_section_start + 1
    hw_out_row = r
    ws_c.cell(row=r, column=1, value="Est HW output (MBH x flow fraction)")
    for s in range(n_scen):
        flow_row = scen_start + s
        formula_cell(ws_c, r, 2 + s, f"=B5*C{flow_row}", dec1_fmt)
    ws_c.cell(row=r, column=2 + n_scen, value="MBH")

    # Simultaneous hours (heating season)
    r += 1
    htg_hrs_row = r
    ws_c.cell(row=r, column=1, value="Simultaneous hours (heating season)")
    formula_cell(ws_c, r, 2, "=Assumptions!B9*Assumptions!B6*B9", num_fmt)
    for s in range(1, n_scen):
        formula_cell(ws_c, r, 2 + s, f"=B{r}", num_fmt)
    ws_c.cell(row=r, column=2 + n_scen, value="hr/yr")

    # Heating waste fraction
    r += 1
    htg_waste_row = r
    ws_c.cell(row=r, column=1, value="Heating waste fraction (fighting CHW cooling)")
    for s in range(n_scen):
        input_cell(ws_c, r, 2 + s, htg_waste_frac, pct_fmt)
    ws_c.cell(row=r, column=2 + n_scen, value=htg_waste_note)

    # Wasted therms
    r += 1
    htg_therms_row = r
    ws_c.cell(row=r, column=1, value="Wasted heating (therms/yr)")
    for s in range(n_scen):
        c = get_column_letter(2 + s)
        formula_cell(ws_c, r, 2 + s, f"={c}{hw_out_row}*{c}{htg_waste_row}*{c}{htg_hrs_row}/100", num_fmt)
    ws_c.cell(row=r, column=2 + n_scen, value="therms/yr")

    # Gas cost
    r += 1
    htg_gas_cost_row = r
    ws_c.cell(row=r, column=1, value="Heating season gas cost")
    for s in range(n_scen):
        c = get_column_letter(2 + s)
        formula_cell(ws_c, r, 2 + s, f"={c}{htg_therms_row}*Assumptions!B4", dollar_fmt)
    ws_c.cell(row=r, column=2 + n_scen, value="$/yr")
    style_range(ws_c, hw_out_row, htg_gas_cost_row, len(htg_headers))

    # ── Cooling Season Electrical Waste ──
    r += 2
    ws_c.cell(row=r, column=1, value="Cooling Season Electrical Waste").font = subsection_font
    clg_section_start = r + 1
    r = clg_section_start
    for i, h in enumerate(htg_headers, 1):
        ws_c.cell(row=r, column=i, value=h)
    style_header_row(ws_c, r, len(htg_headers))

    # Heating load fighting cooling (Btu/hr)
    r += 1
    btu_row = r
    ws_c.cell(row=r, column=1, value="Heating load fighting cooling (Btu/hr)")
    for s in range(n_scen):
        c = get_column_letter(2 + s)
        formula_cell(ws_c, r, 2 + s, f"={c}{hw_out_row}*1000", num_fmt)
    ws_c.cell(row=r, column=2 + n_scen, value="Btu/hr")

    # Chiller kW to remove heat
    r += 1
    chiller_kw_row = r
    ws_c.cell(row=r, column=1, value="Chiller kW to remove heat waste")
    for s in range(n_scen):
        c = get_column_letter(2 + s)
        formula_cell(ws_c, r, 2 + s, f"={c}{btu_row}/(Assumptions!B5*1000)", dec1_fmt)
    ws_c.cell(row=r, column=2 + n_scen, value="kW")

    # Cooling season simultaneous hours
    r += 1
    clg_hrs_row = r
    ws_c.cell(row=r, column=1, value="Cooling season simultaneous hours")
    formula_cell(ws_c, r, 2, "=Assumptions!B8*Assumptions!B6*B9", num_fmt)
    for s in range(1, n_scen):
        formula_cell(ws_c, r, 2 + s, f"=B{r}", num_fmt)
    ws_c.cell(row=r, column=2 + n_scen, value="hr/yr")

    # Wasted chiller kWh
    r += 1
    clg_kwh_row = r
    ws_c.cell(row=r, column=1, value="Wasted chiller kWh")
    for s in range(n_scen):
        c = get_column_letter(2 + s)
        formula_cell(ws_c, r, 2 + s, f"={c}{chiller_kw_row}*{c}{clg_hrs_row}", kwh_fmt)
    ws_c.cell(row=r, column=2 + n_scen, value="kWh/yr")

    # Electricity cost
    r += 1
    clg_elec_cost_row = r
    ws_c.cell(row=r, column=1, value="Wasted electricity cost")
    for s in range(n_scen):
        c = get_column_letter(2 + s)
        formula_cell(ws_c, r, 2 + s, f"={c}{clg_kwh_row}*Assumptions!B3", dollar_fmt)
    ws_c.cell(row=r, column=2 + n_scen, value="$/yr")

    # Gas waste during cooling
    r += 1
    clg_therms_row = r
    ws_c.cell(row=r, column=1, value="Gas waste (heating during cooling season)")
    for s in range(n_scen):
        c = get_column_letter(2 + s)
        formula_cell(ws_c, r, 2 + s, f"={c}{hw_out_row}*{c}{clg_hrs_row}/100", num_fmt)
    ws_c.cell(row=r, column=2 + n_scen, value="therms/yr")

    # Gas cost (cooling season)
    r += 1
    clg_gas_cost_row = r
    ws_c.cell(row=r, column=1, value="Gas cost (cooling season)")
    for s in range(n_scen):
        c = get_column_letter(2 + s)
        formula_cell(ws_c, r, 2 + s, f"={c}{clg_therms_row}*Assumptions!B4", dollar_fmt)
    ws_c.cell(row=r, column=2 + n_scen, value="$/yr")
    style_range(ws_c, clg_section_start + 1, clg_gas_cost_row, len(htg_headers))

    # ── SUMMARY ──
    r += 2
    ws_c.cell(row=r, column=1, value=f"{ecm_id} Summary").font = subsection_font
    sum_start = r + 1
    r = sum_start
    sum_headers = ["Scenario", "Elec kWh Waste", "Elec $/yr", "Gas therms Waste", "Gas $/yr", "Total $/yr"]
    for i, h in enumerate(sum_headers, 1):
        ws_c.cell(row=r, column=i, value=h)
    style_header_row(ws_c, r, 6)

    for s_idx, scen in enumerate(scenarios):
        r = sum_start + 1 + s_idx
        c = get_column_letter(2 + s_idx)
        ws_c.cell(row=r, column=1, value=scen["name"])
        formula_cell(ws_c, r, 2, f"={c}{clg_kwh_row}", kwh_fmt)
        formula_cell(ws_c, r, 3, f"={c}{clg_elec_cost_row}", dollar_fmt)
        formula_cell(ws_c, r, 4, f"={c}{htg_therms_row}+{c}{clg_therms_row}", num_fmt)
        formula_cell(ws_c, r, 5, f"={c}{htg_gas_cost_row}+{c}{clg_gas_cost_row}", dollar_fmt)
        formula_cell(ws_c, r, 6, f"=C{r}+E{r}", dollar_fmt)
    style_range(ws_c, sum_start + 1, sum_start + n_scen, 6)

    # Notes
    r = sum_start + n_scen + 2
    impl_cost = cfg.get("implementation_cost", "")
    if impl_cost:
        ws_c.cell(row=r, column=1, value=f"Implementation cost: {impl_cost}").font = note_font
        r += 1

    verifications = cfg.get("verifications", [])
    if verifications:
        r += 1
        ws_c.cell(row=r, column=1, value="Verification Items").font = subsection_font
        for v_idx, v in enumerate(verifications):
            ws_c.cell(row=r + 1 + v_idx, column=1, value=f"{v_idx + 1}. {v}").font = warn_font

    return wb


def build_config_from_cli(args):
    return {
        "ecm_id": args.ecm_id or "ECM",
        "ecm_title": args.title or "Simultaneous Heating/Cooling",
        "project": args.project or "",
        "elec_rate": args.elec_rate,
        "gas_rate": args.gas_rate,
        "chiller_eer": args.eer,
        "hours_per_week": 168,
        "weeks_per_year": 52,
        "cooling_season_weeks": args.cool_weeks,
        "unit_name": args.name,
        "heating_capacity_mbh": args.htg_mbh,
        "cooling_capacity_tons": args.clg_tons,
        "hw_valve_avg_position": args.hw_valve_pos,
        "chw_valve_avg_position": args.chw_valve_pos,
        "simultaneous_fraction": args.sim_fraction,
        "heating_waste_fraction": args.htg_waste_frac,
        "valve_scenarios": [
            {"name": "Aggressive", "assumed_flow_pct": args.hw_valve_pos, "note": "Position = flow"},
            {"name": "Adjusted", "assumed_flow_pct": 0.15, "note": "Equal-% correction"},
            {"name": "Conservative", "assumed_flow_pct": 0.10, "note": "Equal-% + seasonal"}
        ],
        "verifications": []
    }


def main():
    parser = argparse.ArgumentParser(description="ECM Sim Htg/Clg Savings Calculator — generates .xlsx with Excel formulas")
    parser.add_argument("--config", type=str, help="JSON config file path")
    parser.add_argument("--output", "-o", type=str, default="ecm-simhtgclg-savings.xlsx")
    parser.add_argument("--name", type=str, default="RTU-2")
    parser.add_argument("--htg-mbh", type=float, default=592)
    parser.add_argument("--clg-tons", type=float, default=24)
    parser.add_argument("--hw-valve-pos", type=float, default=0.42)
    parser.add_argument("--chw-valve-pos", type=float, default=0.97)
    parser.add_argument("--sim-fraction", type=float, default=0.48)
    parser.add_argument("--htg-waste-frac", type=float, default=0.10)
    parser.add_argument("--elec-rate", type=float, default=0.10)
    parser.add_argument("--gas-rate", type=float, default=0.90)
    parser.add_argument("--eer", type=float, default=10.5)
    parser.add_argument("--cool-weeks", type=int, default=22)
    parser.add_argument("--ecm-id", type=str, default=None)
    parser.add_argument("--title", type=str, default=None)
    parser.add_argument("--project", type=str, default=None)
    args = parser.parse_args()

    if args.config:
        with open(args.config) as f:
            cfg = json.load(f)
    else:
        cfg = build_config_from_cli(args)

    wb = build_workbook(cfg)
    wb.save(args.output)
    print(f"Saved: {args.output}")


if __name__ == "__main__":
    main()
