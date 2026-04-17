"""ECM Motor Savings Calculator — generates .xlsx with working Excel formulas.

Covers any motor-hour-based ECM:
  - Pump interlocking (pumps running when load equipment is off)
  - Fan scheduling (fans running 24/7 when only needed during occupied hours)
  - Equipment scheduling (general motor runtime reduction)

Usage:
  # CLI args (quick one-off, single equipment)
  py -3.12 ecm-motor-savings.py --name "P-7" --hp 5 --current-hrs-wk 168 --proposed-hrs-wk 84 --elec-rate 0.10 --output savings.xlsx

  # JSON config (reproducible, multi-equipment)
  py -3.12 ecm-motor-savings.py --config inputs.json --output savings.xlsx

JSON config format:
{
  "project": "Example Project",
  "ecm_id": "ECM-04",
  "ecm_title": "CHW Pump Interlock with Chiller & Demand",
  "elec_rate": 0.10,
  "kw_per_hp": 0.746,
  "weeks_per_year": 52,
  "equipment": [
    {
      "name": "P-7 (secondary CHW pump, VFD)",
      "hp": 5,
      "hp_note": "Phase 3.2 estimate; verify nameplate",
      "hp_verified": false,
      "current_hrs_wk": 168,
      "current_note": "Running 24/7 per trend data",
      "proposed_hrs_wk": 84,
      "proposed_note": "Non-cooling season off + 50% cooling season",
      "scenarios": [
        {"name": "Conservative", "hours_saved_wk": 84, "note": "Non-cooling season only"},
        {"name": "Moderate", "hours_saved_wk": 108, "note": "Full year minus 50% cooling"}
      ]
    }
  ],
  "notes": ["Trend data: Jan 28 - Feb 4, 2026"],
  "verifications": ["Confirm P-7 nameplate HP from field"]
}
"""
import argparse, json, sys
from pathlib import Path
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent))
from _styles import *


def build_workbook(cfg):
    wb = Workbook()
    equipment = cfg["equipment"]
    elec_rate = cfg.get("elec_rate", 0.10)
    kw_per_hp = cfg.get("kw_per_hp", 0.746)
    weeks_yr = cfg.get("weeks_per_year", 52)
    ecm_id = cfg.get("ecm_id", "ECM")
    ecm_title = cfg.get("ecm_title", "Motor Savings")
    project = cfg.get("project", "")

    # ── ASSUMPTIONS SHEET ──
    ws = wb.active
    ws.title = "Assumptions"
    set_col_widths(ws, [35, 18, 12, 50])
    section_header(ws, 1, "Assumptions & Constants", 4)
    headers = ["Parameter", "Value", "Units", "Basis / Source"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header_row(ws, 2, 4)

    # Row 3: elec rate
    ws.cell(row=3, column=1, value="Electricity rate")
    input_cell(ws, 3, 2, elec_rate, dollar_fmt2)
    ws.cell(row=3, column=3, value="$/kWh")
    ws.cell(row=3, column=4, value=cfg.get("elec_rate_source", "Verify against utility bills"))

    # Row 4: kW/HP
    ws.cell(row=4, column=1, value="Motor efficiency factor")
    input_cell(ws, 4, 2, kw_per_hp, dec3_fmt)
    ws.cell(row=4, column=3, value="kW/HP")
    ws.cell(row=4, column=4, value="Standard conversion: 1 HP = 0.746 kW")

    # Row 5: weeks/yr
    ws.cell(row=5, column=1, value="Weeks per year")
    input_cell(ws, 5, 2, weeks_yr, num_fmt)
    ws.cell(row=5, column=3, value="wk/yr")

    style_range(ws, 3, 5, 4)
    add_color_legend(ws, 7)
    ws.cell(row=12, column=1, value="All calculation sheets reference this sheet — change an assumption here and all equipment updates automatically.").font = note_font

    # ── EQUIPMENT SHEETS ──
    for eq_idx, equip in enumerate(equipment):
        eq_name = equip["name"]
        sheet_name = eq_name[:31]  # Excel 31-char limit
        ws_eq = wb.create_sheet(sheet_name)
        set_col_widths(ws_eq, [38, 18, 18, 18, 40])

        section_header(ws_eq, 1, f"{ecm_id}: {eq_name}", 5)

        # Equipment info
        r = 3
        ws_eq.cell(row=r, column=1, value="Equipment Data").font = subsection_font

        r = 4
        for i, h in enumerate(["Parameter", "Value", "Units", "Notes"], 1):
            ws_eq.cell(row=r, column=i, value=h)
        style_header_row(ws_eq, r, 4)

        r = 5
        ws_eq.cell(row=r, column=1, value="Motor HP")
        hp_val = equip["hp"]
        if isinstance(hp_val, str):
            # formula like "=822/746"
            formula_cell(ws_eq, r, 2, hp_val, dec3_fmt)
        else:
            input_cell(ws_eq, r, 2, hp_val, dec1_fmt)
        ws_eq.cell(row=r, column=3, value="HP")
        ws_eq.cell(row=r, column=4, value=equip.get("hp_note", "")).font = note_font
        if not equip.get("hp_verified", True):
            ws_eq.cell(row=r, column=4).font = warn_font

        r = 6
        ws_eq.cell(row=r, column=1, value="Current runtime")
        input_cell(ws_eq, r, 2, equip["current_hrs_wk"], num_fmt)
        ws_eq.cell(row=r, column=3, value="hr/wk")
        ws_eq.cell(row=r, column=4, value=equip.get("current_note", "")).font = note_font
        style_range(ws_eq, 5, 6, 4)

        # Scenarios
        scenarios = equip.get("scenarios", [])
        if not scenarios:
            proposed = equip.get("proposed_hrs_wk", equip["current_hrs_wk"])
            scenarios = [{"name": "Default", "hours_saved_wk": equip["current_hrs_wk"] - proposed, "note": equip.get("proposed_note", "")}]

        r = 8
        ws_eq.cell(row=r, column=1, value="Runtime Reduction Scenarios").font = subsection_font

        r = 9
        for i, h in enumerate(["Scenario", "Hours Saved/wk", "Hours Saved/yr", "Notes"], 1):
            ws_eq.cell(row=r, column=i, value=h)
        style_header_row(ws_eq, r, 4)

        scenario_start = 10
        for s_idx, scenario in enumerate(scenarios):
            r = scenario_start + s_idx
            ws_eq.cell(row=r, column=1, value=scenario["name"])
            input_cell(ws_eq, r, 2, scenario["hours_saved_wk"], num_fmt)
            # Hours saved/yr = hours_saved_wk * weeks_per_year
            formula_cell(ws_eq, r, 3, f"=B{r}*Assumptions!B5", num_fmt)
            ws_eq.cell(row=r, column=4, value=scenario.get("note", "")).font = note_font
        style_range(ws_eq, scenario_start, scenario_start + len(scenarios) - 1, 4)

        # Savings calculations
        calc_start = scenario_start + len(scenarios) + 2
        r = calc_start
        ws_eq.cell(row=r, column=1, value="Savings Calculations").font = subsection_font

        r = calc_start + 1
        calc_headers = ["Calculation Step"] + [s["name"] for s in scenarios] + ["Units"]
        n_scen = len(scenarios)
        for i, h in enumerate(calc_headers, 1):
            ws_eq.cell(row=r, column=i, value=h)
        style_header_row(ws_eq, r, len(calc_headers))

        # kW draw
        r = calc_start + 2
        ws_eq.cell(row=r, column=1, value="Motor kW (HP x kW/HP)")
        for s in range(n_scen):
            formula_cell(ws_eq, r, 2 + s, f"=B5*Assumptions!B4", dec3_fmt)
        ws_eq.cell(row=r, column=2 + n_scen, value="kW")
        kw_row = r

        # Annual kWh saved per scenario
        r = calc_start + 3
        ws_eq.cell(row=r, column=1, value="Annual kWh saved")
        for s in range(n_scen):
            scen_col = get_column_letter(2 + s)
            hrs_row = scenario_start + s
            formula_cell(ws_eq, r, 2 + s, f"={scen_col}{kw_row}*C{hrs_row}", kwh_fmt)
        ws_eq.cell(row=r, column=2 + n_scen, value="kWh/yr")
        kwh_row = r

        # Annual $ saved
        r = calc_start + 4
        ws_eq.cell(row=r, column=1, value="Annual $ saved")
        for s in range(n_scen):
            scen_col = get_column_letter(2 + s)
            formula_cell(ws_eq, r, 2 + s, f"={scen_col}{kwh_row}*Assumptions!B3", dollar_fmt)
        ws_eq.cell(row=r, column=2 + n_scen, value="$/yr")
        dollar_row = r
        style_range(ws_eq, calc_start + 2, calc_start + 4, len(calc_headers))

        # Store row references for summary
        equip["_sheet"] = sheet_name
        equip["_kwh_row"] = kwh_row
        equip["_dollar_row"] = dollar_row
        equip["_n_scen"] = n_scen
        equip["_scenario_names"] = [s["name"] for s in scenarios]

        # Notes
        r = calc_start + 6
        if not equip.get("hp_verified", True):
            ws_eq.cell(row=r, column=1, value=f"VERIFICATION NEEDED: Confirm {eq_name} nameplate HP from field.").font = warn_font
            r += 1
        impl_cost = equip.get("implementation_cost", "")
        if impl_cost:
            ws_eq.cell(row=r, column=1, value=f"Implementation cost: {impl_cost}").font = note_font

    # ── SUMMARY SHEET ──
    ws_s = wb.create_sheet("Summary")
    set_col_widths(ws_s, [8, 35, 18, 18, 14, 40])
    section_header(ws_s, 1, f"{ecm_id}: {ecm_title} — Summary", 6)
    if project:
        ws_s.cell(row=2, column=1, value=f"Project: {project}").font = note_font
    ws_s.cell(row=3, column=1, value="All values reference calculation sheets via formulas.").font = note_font

    # Determine max scenarios across all equipment
    max_scen = max(eq.get("_n_scen", 1) for eq in equipment)
    # Use first equipment's scenario names as column headers (typical case: all same)
    scen_names = equipment[0].get("_scenario_names", ["Default"])

    r = 5
    sum_headers = ["#", "Equipment"] + [f"kWh/yr ({s})" for s in scen_names] + [f"$/yr ({s})" for s in scen_names]
    for i, h in enumerate(sum_headers, 1):
        ws_s.cell(row=r, column=i, value=h)
    style_header_row(ws_s, r, len(sum_headers))

    data_start = 6
    for eq_idx, equip in enumerate(equipment):
        r = data_start + eq_idx
        ws_s.cell(row=r, column=1, value=eq_idx + 1)
        ws_s.cell(row=r, column=2, value=equip["name"])
        n_s = equip.get("_n_scen", 1)
        sheet = equip["_sheet"]
        kwh_r = equip["_kwh_row"]
        dol_r = equip["_dollar_row"]
        for s in range(min(n_s, max_scen)):
            col_letter = get_column_letter(2 + s)
            green_cell(ws_s, r, 3 + s, f"='{sheet}'!{col_letter}{kwh_r}", kwh_fmt)
            green_cell(ws_s, r, 3 + max_scen + s, f"='{sheet}'!{col_letter}{dol_r}", dollar_fmt)
    style_range(ws_s, data_start, data_start + len(equipment) - 1, len(sum_headers))

    # Totals row
    r = data_start + len(equipment)
    ws_s.cell(row=r, column=2, value="TOTAL").font = black_bold
    for s in range(max_scen):
        kwh_col = get_column_letter(3 + s)
        dol_col = get_column_letter(3 + max_scen + s)
        formula_cell(ws_s, r, 3 + s, f"=SUM({kwh_col}{data_start}:{kwh_col}{r - 1})", kwh_fmt)
        ws_s.cell(row=r, column=3 + s).font = Font(bold=True)
        formula_cell(ws_s, r, 3 + max_scen + s, f"=SUM({dol_col}{data_start}:{dol_col}{r - 1})", dollar_fmt)
        ws_s.cell(row=r, column=3 + max_scen + s).font = Font(bold=True)
    style_range(ws_s, r, r, len(sum_headers))

    # Verifications
    verifications = cfg.get("verifications", [])
    if verifications:
        r = data_start + len(equipment) + 2
        ws_s.cell(row=r, column=1, value="Verification Items").font = subsection_font
        for v_idx, v in enumerate(verifications):
            ws_s.cell(row=r + 1 + v_idx, column=1, value=f"{v_idx + 1}.").font = warn_font
            ws_s.cell(row=r + 1 + v_idx, column=2, value=v).font = warn_font

    return wb


def build_config_from_cli(args):
    return {
        "ecm_id": args.ecm_id or "ECM",
        "ecm_title": args.title or "Motor Savings",
        "project": args.project or "",
        "elec_rate": args.elec_rate,
        "kw_per_hp": args.kw_per_hp,
        "weeks_per_year": args.weeks_per_year,
        "equipment": [{
            "name": args.name,
            "hp": args.hp,
            "hp_verified": not args.unverified,
            "current_hrs_wk": args.current_hrs_wk,
            "proposed_hrs_wk": args.proposed_hrs_wk,
            "scenarios": [
                {"name": "Single Scenario", "hours_saved_wk": args.current_hrs_wk - args.proposed_hrs_wk}
            ]
        }],
        "verifications": []
    }


def main():
    parser = argparse.ArgumentParser(description="ECM Motor Savings Calculator — generates .xlsx with Excel formulas")
    parser.add_argument("--config", type=str, help="JSON config file path")
    parser.add_argument("--output", "-o", type=str, default="ecm-motor-savings.xlsx")
    # CLI-only args for single equipment
    parser.add_argument("--name", type=str, default="Motor-1")
    parser.add_argument("--hp", type=float, default=5.0)
    parser.add_argument("--current-hrs-wk", type=float, default=168)
    parser.add_argument("--proposed-hrs-wk", type=float, default=84)
    parser.add_argument("--elec-rate", type=float, default=0.10)
    parser.add_argument("--kw-per-hp", type=float, default=0.746)
    parser.add_argument("--weeks-per-year", type=int, default=52)
    parser.add_argument("--ecm-id", type=str, default=None)
    parser.add_argument("--title", type=str, default=None)
    parser.add_argument("--project", type=str, default=None)
    parser.add_argument("--unverified", action="store_true", help="Flag HP as unverified")
    args = parser.parse_args()

    if args.config:
        with open(args.config) as f:
            cfg = json.load(f)
    else:
        cfg = build_config_from_cli(args)

    wb = build_workbook(cfg)
    wb.save(args.output)
    print(f"Saved: {args.output}")


if __name__ == "__main__":
    main()
