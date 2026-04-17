#!/usr/bin/env python3
"""
AHU Fault Detection Tool - Reusable RCx Analysis

Analyzes AHU trend data to identify common operational faults and generates
Excel workbook with findings, calculations, and visualization charts.

Usage:
    python ahu-fault-detector.py --csv <path_to_trends.csv> --output <output.xlsx> [options]

Example:
    python ahu-fault-detector.py \
        --csv ahu1-trends.csv \
        --output ahu1-analysis.xlsx \
        --oa-temp-col "AHU-1 : OSA Temp (°F)" \
        --damper-col "AHU-1 : Economizer Dmpr OA/RA Status (%)" \
        --cooling-tons 10 \
        --elec-rate 0.10

Author: Mat Coalson
Created: 2026-02-13
"""

import pandas as pd
import numpy as np
import argparse
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.utils.dataframe import dataframe_to_rows


def create_excel_output(df, faults, args, output_path):
    """Generate Excel workbook with findings, data, calculations, and charts."""

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: Summary
    ws_summary = wb.create_sheet("Summary", 0)
    create_summary_sheet(ws_summary, faults, args)

    # Sheet 2: Calculations (shows the math)
    ws_calc = wb.create_sheet("Calculations", 1)
    create_calculations_sheet(ws_calc, df, faults, args)

    # Sheet 3: Trend Data
    ws_data = wb.create_sheet("Trend Data", 2)
    create_data_sheet(ws_data, df, args)

    # Sheet 4: Charts
    ws_charts = wb.create_sheet("Visualizations", 3)
    create_charts_sheet(ws_charts, df, args)

    wb.save(output_path)


def create_summary_sheet(ws, faults, args):
    """Summary of findings and recommendations."""

    # Title
    ws['A1'] = 'AHU FAULT DETECTION ANALYSIS'
    ws['A1'].font = Font(size=16, bold=True)

    ws['A2'] = f'Source File: {args.csv}'
    ws['A3'] = f'Analysis Date: {pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")}'

    # Findings header
    ws['A5'] = 'FAULTS DETECTED'
    ws['A5'].font = Font(size=14, bold=True)
    ws['A5'].fill = PatternFill('solid', start_color='D3D3D3')

    row = 7
    for i, fault in enumerate(faults, 1):
        # Fault number and type
        ws[f'A{row}'] = f"{i}. {fault['fault_type'].upper().replace('_', ' ')}"
        ws[f'A{row}'].font = Font(bold=True, size=12)

        # Severity
        ws[f'B{row}'] = f"Severity: {fault['severity'].upper()}"
        severity_color = {'high': 'FF0000', 'medium': 'FFA500', 'low': 'FFFF00'}
        ws[f'B{row}'].fill = PatternFill('solid', start_color=severity_color.get(fault['severity'], 'FFFFFF'))

        row += 1

        # Finding
        ws[f'A{row}'] = 'Finding:'
        ws[f'A{row}'].font = Font(italic=True)
        ws[f'B{row}'] = fault['finding']
        row += 1

        # Recommendation
        ws[f'A{row}'] = 'Recommendation:'
        ws[f'A{row}'].font = Font(italic=True)
        ws[f'B{row}'] = fault['recommendation']
        row += 1

        # Savings if applicable
        if 'annual_savings_dollars' in fault:
            ws[f'A{row}'] = 'Annual Savings Opportunity:'
            ws[f'A{row}'].font = Font(italic=True)
            ws[f'B{row}'] = f"${fault['annual_savings_dollars']:,.0f}/year ({fault['annual_savings_kwh']:,.0f} kWh)"
            ws[f'B{row}'].font = Font(color='008000')
            row += 1

        row += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 80


def create_calculations_sheet(ws, df, faults, args):
    """Show the math - auditable calculations with Excel formulas."""

    ws['A1'] = 'CALCULATIONS - TRANSPARENT & AUDITABLE'
    ws['A1'].font = Font(size=14, bold=True)

    ws['A3'] = 'ASSUMPTIONS (User Inputs)'
    ws['A3'].font = Font(size=12, bold=True)
    ws['A3'].fill = PatternFill('solid', start_color='FFFF00')

    # Key assumptions
    row = 4
    assumptions = [
        ('Cooling Capacity (tons)', args.cooling_tons),
        ('Electricity Rate ($/kWh)', args.elec_rate),
        ('kW per Ton (chiller efficiency)', 1.0),
        ('Data Interval (minutes)', 15),
        ('Days in Sample Period', len(df) * 15 / 60 / 24),
    ]

    for label, value in assumptions:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'B{row}'].font = Font(color='0000FF')  # Blue = user input
        ws[f'B{row}'].fill = PatternFill('solid', start_color='FFFF00')
        row += 1

    row += 2

    # Economizer fault calculations
    for fault in faults:
        if fault['fault_type'] == 'economizer_stuck':
            ws[f'A{row}'] = 'ECONOMIZER FAULT - ENERGY SAVINGS CALCULATION'
            ws[f'A{row}'].font = Font(size=12, bold=True)
            row += 1

            ws[f'A{row}'] = 'Free Cooling Hours (from data)'
            ws[f'B{row}'] = fault['hours_affected']
            ws[f'B{row}'].font = Font(color='0000FF')
            row += 1

            ws[f'A{row}'] = 'Annualization Factor (365/sample days)'
            ws[f'B{row}'] = f'=365/B8'
            row += 1

            ws[f'A{row}'] = 'Annual Free Cooling Hours'
            ws[f'B{row}'] = f'=B{row-1}*B{row-2}'
            row += 1

            ws[f'A{row}'] = 'Cooling Load (tons)'
            ws[f'B{row}'] = '=B4'
            ws[f'B{row}'].font = Font(color='00FF00')  # Green = link to assumption
            row += 1

            ws[f'A{row}'] = 'kW per Ton'
            ws[f'B{row}'] = '=B6'
            ws[f'B{row}'].font = Font(color='00FF00')
            row += 1

            ws[f'A{row}'] = 'Free Cooling Displacement Factor'
            ws[f'B{row}'] = 0.7
            ws[f'B{row}'].font = Font(color='0000FF')
            row += 1

            ws[f'A{row}'] = 'Annual kWh Savings'
            ws[f'B{row}'] = f'=B{row-4}*B{row-3}*B{row-2}*B{row-1}'
            ws[f'B{row}'].font = Font(bold=True)
            row += 1

            ws[f'A{row}'] = 'Electricity Rate ($/kWh)'
            ws[f'B{row}'] = '=B5'
            ws[f'B{row}'].font = Font(color='00FF00')
            row += 1

            ws[f'A{row}'] = 'Annual Savings ($)'
            ws[f'B{row}'] = f'=B{row-2}*B{row-1}'
            ws[f'B{row}'].font = Font(bold=True, color='008000')
            ws[f'B{row}'].number_format = '$#,##0'

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20


def create_data_sheet(ws, df, args):
    """Raw trend data for reference."""

    ws['A1'] = 'TREND DATA'
    ws['A1'].font = Font(size=14, bold=True)

    # Write column headers
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(3, col_idx, col_name)
        ws.cell(3, col_idx).font = Font(bold=True)

    # Write data (limit to key columns for demo)
    key_cols = ['Datetime', args.oa_temp_col, args.damper_col,
                args.hw_valve_col, args.chw_valve_col]

    df_subset = df[key_cols].copy()

    for row_idx, row_data in enumerate(dataframe_to_rows(df_subset, index=False, header=False), start=4):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row_idx, col_idx, value)


def create_charts_sheet(ws, df, args):
    """Visualization charts showing the faults."""

    ws['A1'] = 'FAULT VISUALIZATION'
    ws['A1'].font = Font(size=14, bold=True)

    # Chart 1: Economizer Position vs OA Temp
    ws['A3'] = 'Chart 1: Economizer Stuck at Minimum During Free Cooling Opportunity'
    ws['A3'].font = Font(size=11, bold=True)

    # Prepare data for chart
    chart_data = df[[args.oa_temp_col, args.damper_col]].copy()
    chart_data.columns = ['OA Temp', 'Damper Position']

    # Write chart data starting at A5
    ws['A5'] = 'OA Temp (°F)'
    ws['B5'] = 'Damper Position (%)'
    ws['A5'].font = Font(bold=True)
    ws['B5'].font = Font(bold=True)

    for idx, (oa_temp, damper) in enumerate(zip(chart_data['OA Temp'][:50], chart_data['Damper Position'][:50]), start=6):
        ws[f'A{idx}'] = oa_temp
        ws[f'B{idx}'] = damper

    # Create line chart
    chart1 = LineChart()
    chart1.title = "Economizer Damper vs Outside Air Temperature"
    chart1.y_axis.title = 'Value'
    chart1.x_axis.title = 'Time'

    data = Reference(ws, min_col=2, min_row=5, max_row=55)
    cats = Reference(ws, min_col=1, min_row=6, max_row=55)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)

    ws.add_chart(chart1, "D5")

    # Chart 2: Simultaneous Heating/Cooling
    ws['A30'] = 'Chart 2: Simultaneous Heating and Cooling (Fighting Valves)'
    ws['A30'].font = Font(size=11, bold=True)

    ws['A32'] = 'HW Valve (%)'
    ws['B32'] = 'CHW Valve (%)'
    ws['A32'].font = Font(bold=True)
    ws['B32'].font = Font(bold=True)

    for idx, (hw, chw) in enumerate(zip(df[args.hw_valve_col][:50], df[args.chw_valve_col][:50]), start=33):
        ws[f'A{idx}'] = hw
        ws[f'B{idx}'] = chw

    chart2 = LineChart()
    chart2.title = "HW vs CHW Valve Position (Conflict Detection)"
    chart2.y_axis.title = 'Valve Position (%)'
    chart2.x_axis.title = 'Time'

    data2 = Reference(ws, min_col=1, min_row=32, max_col=2, max_row=82)
    chart2.add_data(data2, titles_from_data=True)

    ws.add_chart(chart2, "D30")


def analyze_economizer(df, oa_temp_col, damper_col, occ_col, cooling_tons, elec_rate):
    """Detect economizer stuck at minimum position during free cooling opportunities."""

    free_cooling = df[
        (df[oa_temp_col] >= 55) &
        (df[oa_temp_col] <= 65) &
        (df[occ_col] == 'Occupied')
    ]

    if len(free_cooling) == 0:
        return None

    avg_damper = free_cooling[damper_col].mean()
    hours = len(free_cooling) * 15 / 60

    fault = {
        'fault_type': 'economizer_stuck',
        'severity': 'medium' if avg_damper < 20 else 'low',
        'hours_affected': hours,
        'avg_damper_position': avg_damper,
        'expected_position': '80-100%',
    }

    if avg_damper < 20:
        annual_hours = hours * (365 / 30)
        annual_kwh = cooling_tons * 1.0 * annual_hours * 0.7
        annual_cost = annual_kwh * elec_rate

        fault['annual_savings_kwh'] = round(annual_kwh, 0)
        fault['annual_savings_dollars'] = round(annual_cost, 0)
        fault['finding'] = (
            f"Economizer damper stuck at {avg_damper:.0f}% during {hours:.0f} hours "
            f"of free cooling opportunity. Annual savings potential: ${annual_cost:,.0f}."
        )
        fault['recommendation'] = (
            "Inspect economizer damper actuator and linkage. "
            "Verify control sequence in BAS. "
            "Test damper through full stroke. "
            "Commission per ASHRAE 90.1."
        )

    return fault


def analyze_simultaneous_htg_clg(df, hw_col, chw_col):
    """Detect simultaneous heating and cooling valve operation."""

    simul = df[(df[hw_col] > 5) & (df[chw_col] > 5)]

    if len(simul) == 0:
        return None

    hours = len(simul) * 15 / 60
    avg_hw = simul[hw_col].mean()
    avg_chw = simul[chw_col].mean()

    fault = {
        'fault_type': 'simultaneous_heating_cooling',
        'severity': 'high' if avg_hw > 50 and avg_chw > 50 else 'medium',
        'hours_affected': hours,
        'avg_hw_valve': avg_hw,
        'avg_chw_valve': avg_chw,
        'finding': (
            f"Simultaneous heating/cooling detected for {hours:.0f} hours. "
            f"HW valve avg {avg_hw:.0f}%, CHW valve avg {avg_chw:.0f}%. "
            f"Fighting control valves waste heating and cooling energy."
        ),
        'recommendation': (
            "Review AHU control sequence for proper valve staging. "
            "Implement dead-band between heating and cooling modes. "
            "Check for sensor calibration issues. "
            "Consider interlock to prevent simultaneous operation."
        )
    }

    return fault


def main():
    parser = argparse.ArgumentParser(description='AHU fault detection with Excel output')
    parser.add_argument('--csv', type=str, required=True)
    parser.add_argument('--output', type=str, required=True)
    parser.add_argument('--oa-temp-col', type=str, default='OSA Temp (°F)')
    parser.add_argument('--damper-col', type=str, default='Economizer Dmpr Status (%)')
    parser.add_argument('--hw-valve-col', type=str, default='Hot Water Valve Status (%)')
    parser.add_argument('--chw-valve-col', type=str, default='Chilled Water Valve Status (%)')
    parser.add_argument('--occ-col', type=str, default='Occupancy Status')
    parser.add_argument('--supply-fan-col', type=str, default='Supply Fan VFD Signal Status (%)')
    parser.add_argument('--cooling-tons', type=float, default=10.0)
    parser.add_argument('--elec-rate', type=float, default=0.10)

    args = parser.parse_args()

    print("=" * 80)
    print("AHU FAULT DETECTION TOOL")
    print("=" * 80)
    print(f"Reading: {args.csv}")

    df = pd.read_csv(args.csv)
    df['Datetime'] = pd.to_datetime(df.iloc[:, 0])

    print(f"Dataset: {len(df)} points, {df['Datetime'].min()} to {df['Datetime'].max()}")
    print()

    faults = []

    if args.oa_temp_col in df.columns and args.damper_col in df.columns:
        fault = analyze_economizer(
            df, args.oa_temp_col, args.damper_col, args.occ_col,
            args.cooling_tons, args.elec_rate
        )
        if fault:
            faults.append(fault)

    if args.hw_valve_col in df.columns and args.chw_valve_col in df.columns:
        fault = analyze_simultaneous_htg_clg(df, args.hw_valve_col, args.chw_valve_col)
        if fault:
            faults.append(fault)

    print(f"FAULTS DETECTED: {len(faults)}")
    for i, fault in enumerate(faults, 1):
        print(f"{i}. {fault['fault_type'].upper()}: {fault['finding']}")
    print()

    # Generate Excel output
    print(f"Generating Excel workbook: {args.output}")
    create_excel_output(df, faults, args, args.output)
    print(f"✓ Excel workbook created with Summary, Calculations, Data, and Charts")

    return faults


if __name__ == '__main__':
    main()
