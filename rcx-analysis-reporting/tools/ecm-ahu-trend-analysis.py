#!/usr/bin/env python3
"""
AHU Trend Data Analysis Tool — Reusable RCx Analysis

Analyzes AHU trend data from BAS exports, detects operational faults,
quantifies energy waste, and generates an Excel workbook with:
  - Summary sheet with all findings and recommendations
  - Calculations sheet with working Excel formulas (auditable)
  - Fault Data sheets with flagged intervals per fault
  - Charts showing fault signatures

Fault detection rules:
  1. Economizer stuck/locked — damper fails to modulate during free-cooling conditions
  2. HW valve stuck open — heating valve pegged with no modulation
  3. Simultaneous heating/cooling — HW and CHW valves open concurrently
  4. DAT not tracking setpoint — discharge air temp consistently above setpoint
  5. Humidification offline — steam system non-functional, humidity far below setpoint
  6. Supply fan at max speed — VFD not modulating (possible duct static issue)

Usage:
    python ecm-ahu-trend-analysis.py --csv <trends.csv> --output <output.xlsx> [options]
    python ecm-ahu-trend-analysis.py --config <config.json> --output <output.xlsx>

Author: Mat Coalson
Created: 2026-02-22
"""

import pandas as pd
import numpy as np
import argparse
import json
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter

# Import shared styles
sys.path.insert(0, str(Path(__file__).parent))
from _styles import (
    blue_font, blue_bold, black_bold, green_font, header_font,
    section_font, subsection_font, note_font, warn_font,
    header_fill, assumption_fill, section_fill, thin_border,
    pct_fmt, kwh_fmt, dollar_fmt, dollar_fmt2, num_fmt, dec1_fmt, dec2_fmt,
    style_header_row, style_range, set_col_widths,
    input_cell, formula_cell, green_cell, section_header, add_color_legend
)
from openpyxl.styles import Font, PatternFill, Alignment


# ── Column mapping defaults (example AHU) ──────────────────────────
DEFAULT_COLS = {
    "datetime":    "Datetime",
    "occ":         "AHU-1 : Occupancy Status",
    "econ_dmpr":   "AHU-1 : Economizer Dmpr OA/RA Status (%)",
    "hw_valve":    "AHU-1 : Hot Water Valve Status (%)",
    "chw_valve":   "AHU-1 : Chilled Water Valve Status (%)",
    "oat":         "AHU-1 : OSA Temp (°F)",
    "mat":         "AHU-1 : Mixed Air Temp Status (°F)",
    "dat":         "AHU-1 : Discharge Air Temp Status (°F)",
    "dat_sp":      "AHU-1 : DAT Setpoint Active (°F)",
    "rat":         "AHU-1 : Return Air Temp Status (°F)",
    "sf_spd":      "AHU-1 : Supply Fan VFD Signal Status (%)",
    "rf_spd":      "AHU-1 : Return Fan VFD Signal Status (%)",
    "sf_status":   "AHU-1 : Supply Fan Status",
    "stm_press":   "AHU-1 : Steam Pressure (psi)",
    "humidity":    "AHU-1 : AHU-1 Average Humidity (%)",
    "hum_sp":      "AHU-1 : AHU-1 Spce Humidity Stpt (%)",
    "sih1":        "AHU-1 : SIH-1 Signal Status (%)",
    "sih2":        "AHU-1 : SIH-2 Signal Status (%)",
    "sih3":        "AHU-1 : SIH-3 Signal Status (%)",
}


# ── Fault Detection Functions ─────────────────────────────────────────────────

def detect_economizer_stuck(df, cols, params):
    """Detect economizer damper locked at minimum during free-cooling opportunity."""
    econ = df[cols["econ_dmpr"]]
    oat = df[cols["oat"]]
    n_unique = econ.nunique()
    min_pos = econ.min()
    max_pos = econ.max()

    # Free-cooling window: OAT between low and high economizer thresholds
    lo = params.get("econ_lo_f", 45)
    hi = params.get("econ_hi_f", 70)
    fc_mask = (oat >= lo) & (oat <= hi)
    fc_rows = df[fc_mask]
    fc_hours = len(fc_rows) * params["interval_min"] / 60

    stuck = n_unique <= 3 and (max_pos - min_pos) < 5
    if not stuck:
        return None

    return {
        "id": "ECON-STUCK",
        "title": "Economizer Damper Locked at Minimum Position",
        "severity": "HIGH",
        "finding": (
            f"Economizer damper locked at {min_pos:.0f}% for the entire {params['sample_days']:.0f}-day "
            f"sample period with zero modulation ({n_unique} unique value(s)). "
            f"Free-cooling conditions (OAT {lo}–{hi}°F) occurred for {fc_hours:.0f} hours "
            f"({100*len(fc_rows)/len(df):.0f}% of dataset). During these periods the damper "
            f"should have modulated toward 100% to displace mechanical cooling."
        ),
        "recommendation": (
            "Inspect economizer damper actuator, linkage, and end switches. "
            "Verify BAS economizer enable/disable logic and changeover setpoint. "
            "Test damper through full stroke from BAS. "
            "Check for actuator power and control signal at the damper. "
            "Possible causes: failed actuator, disconnected linkage, BAS logic error, "
            "or wiring/communication fault."
        ),
        "data": {
            "damper_position_pct": float(min_pos),
            "unique_positions": int(n_unique),
            "free_cooling_hours": round(fc_hours, 1),
            "free_cooling_pct": round(100 * len(fc_rows) / len(df), 1),
        },
        "fc_mask": fc_mask,
    }


def detect_hw_valve_stuck(df, cols, params):
    """Detect heating valve stuck open regardless of conditions."""
    hw = df[cols["hw_valve"]]
    oat = df[cols["oat"]]
    n_unique = hw.nunique()
    min_pos = hw.min()
    mean_pos = hw.mean()

    # Check if HW valve is pegged high even when OAT is warm
    warm_mask = oat > 65
    warm_hw = hw[warm_mask]

    if min_pos < 80 or mean_pos < 90:
        return None

    warm_hours = len(df[warm_mask]) * params["interval_min"] / 60
    warm_avg_hw = warm_hw.mean() if len(warm_hw) > 0 else 0

    return {
        "id": "HW-STUCK",
        "title": "Hot Water Valve Locked Open",
        "severity": "CRITICAL",
        "finding": (
            f"Hot water valve stuck at {min_pos:.0f}–{hw.max():.0f}% (mean {mean_pos:.1f}%) "
            f"for the entire sample period. The valve never modulated below {min_pos:.0f}%, "
            f"even during {warm_hours:.0f} hours when OAT exceeded 65°F (avg HW position during "
            f"warm periods: {warm_avg_hw:.0f}%). This forces the chilled water valve to fight "
            f"the heating valve to maintain discharge air temperature."
        ),
        "recommendation": (
            "Inspect HW valve actuator and verify correct control signal. "
            "Check for failed actuator (spring-return to open), disconnected linkage, "
            "or stuck valve stem. Verify BAS heating PID loop output signal corresponds "
            "to actual valve position. Test valve through full stroke from BAS. "
            "Possible causes: failed actuator, frozen valve stem, BAS output error."
        ),
        "data": {
            "min_position_pct": float(min_pos),
            "max_position_pct": float(hw.max()),
            "mean_position_pct": round(float(mean_pos), 1),
            "warm_period_hours": round(warm_hours, 1),
            "warm_period_avg_hw": round(float(warm_avg_hw), 1),
        },
    }


def detect_simul_htg_clg(df, cols, params):
    """Detect simultaneous heating and cooling."""
    hw = df[cols["hw_valve"]]
    chw = df[cols["chw_valve"]]
    threshold = params.get("simul_threshold_pct", 5)

    simul_mask = (hw > threshold) & (chw > threshold)
    simul = df[simul_mask]
    if len(simul) == 0:
        return None

    hours = len(simul) * params["interval_min"] / 60
    avg_hw = hw[simul_mask].mean()
    avg_chw = chw[simul_mask].mean()

    return {
        "id": "SIMUL-HTG-CLG",
        "title": "Simultaneous Heating and Cooling",
        "severity": "HIGH",
        "finding": (
            f"Simultaneous heating and cooling detected for {hours:.0f} hours "
            f"({100*len(simul)/len(df):.0f}% of dataset). During these periods the HW valve "
            f"averaged {avg_hw:.0f}% and CHW valve averaged {avg_chw:.0f}%. "
            f"With HW locked near 100%, CHW must open to counteract excess heating — "
            f"wasting both heating and cooling energy."
        ),
        "recommendation": (
            "Root cause is likely the stuck HW valve (see HW-STUCK finding). "
            "Resolving the HW valve fault should eliminate this simultaneous operation. "
            "After HW valve repair, verify proper dead-band between heating and cooling "
            "modes in the BAS sequence. Minimum 2°F dead-band recommended."
        ),
        "data": {
            "simul_hours": round(hours, 1),
            "simul_pct": round(100 * len(simul) / len(df), 1),
            "avg_hw_pct": round(float(avg_hw), 1),
            "avg_chw_pct": round(float(avg_chw), 1),
            "intervals_affected": int(len(simul)),
        },
        "simul_mask": simul_mask,
    }


def detect_dat_tracking(df, cols, params):
    """Detect discharge air temp not tracking setpoint."""
    dat = df[cols["dat"]]
    dat_sp = df[cols["dat_sp"]]
    err = dat - dat_sp
    threshold = params.get("dat_err_threshold_f", 5)

    over_mask = err > threshold
    over_count = over_mask.sum()
    if over_count < len(df) * 0.1:
        return None

    hours = over_count * params["interval_min"] / 60

    return {
        "id": "DAT-TRACKING",
        "title": "Discharge Air Temp Not Tracking Setpoint",
        "severity": "HIGH",
        "finding": (
            f"Discharge air temperature exceeded setpoint by more than {threshold}°F for "
            f"{hours:.0f} hours ({100*over_count/len(df):.0f}% of dataset). "
            f"Mean DAT error: +{err.mean():.1f}°F. Max overshoot: +{err.max():.1f}°F. "
            f"This is consistent with the stuck HW valve forcing excess reheat."
        ),
        "recommendation": (
            "Root cause is likely the stuck HW valve (see HW-STUCK finding). "
            "After HW valve repair, verify DAT PID loop tuning. Check DAT sensor "
            "calibration against a reference thermometer. Confirm proper sensor location "
            "in the discharge plenum (not too close to coils)."
        ),
        "data": {
            "over_threshold_hours": round(hours, 1),
            "over_threshold_pct": round(100 * over_count / len(df), 1),
            "mean_error_f": round(float(err.mean()), 1),
            "max_error_f": round(float(err.max()), 1),
        },
        "over_mask": over_mask,
    }


def detect_humidification_offline(df, cols, params):
    """Detect steam humidification system completely non-functional."""
    hum = df[cols["humidity"]]
    hum_sp = df[cols["hum_sp"]]
    stm = df[cols["stm_press"]]

    # Check if all humidifier signals are zero
    sih_cols = [c for k, c in cols.items() if k.startswith("sih") and c in df.columns]
    all_zero = all(df[c].max() == 0 for c in sih_cols)
    stm_bad = stm.max() < 0  # negative = sensor fault or no steam

    if not all_zero:
        return None

    below_sp = hum < hum_sp
    below_hours = below_sp.sum() * params["interval_min"] / 60

    return {
        "id": "HUMIDIFICATION-OFFLINE",
        "title": "Steam Humidification System Non-Functional",
        "severity": "CRITICAL" if params.get("museum", False) else "HIGH",
        "finding": (
            f"All three steam injection humidifiers (SIH-1, SIH-2, SIH-3) show 0% signal "
            f"for the entire sample period. Steam pressure reads {stm.iloc[0]:.1f} psi "
            f"({'negative — likely sensor fault or no steam supply' if stm_bad else 'low pressure'}). "
            f"Average space humidity is {hum.mean():.1f}% vs setpoint of {hum_sp.iloc[0]:.0f}%. "
            f"Humidity was below setpoint for {below_hours:.0f} hours "
            f"({100*below_sp.sum()/len(df):.0f}% of dataset)."
        ),
        "recommendation": (
            "Verify steam supply to humidifiers — check boiler/steam generator status, "
            "isolation valves, and steam trap operation. Inspect steam pressure sensor "
            "for calibration (negative reading indicates fault). Check BAS humidification "
            "enable logic and control signals to SIH valves. "
            "For museum applications, sustained low humidity poses risk to artifact preservation."
        ),
        "data": {
            "avg_humidity_pct": round(float(hum.mean()), 1),
            "humidity_sp_pct": float(hum_sp.iloc[0]),
            "steam_pressure_psi": round(float(stm.iloc[0]), 1),
            "below_sp_hours": round(below_hours, 1),
            "below_sp_pct": round(100 * below_sp.sum() / len(df), 1),
        },
    }


def detect_fan_at_max(df, cols, params):
    """Detect supply fan running at maximum speed with minimal modulation."""
    sf = df[cols["sf_spd"]]
    at_max = sf >= params.get("fan_max_threshold_pct", 99)
    pct_at_max = 100 * at_max.sum() / len(df)

    if pct_at_max < 30:
        return None

    return {
        "id": "FAN-AT-MAX",
        "title": "Supply Fan Running at Maximum Speed",
        "severity": "MEDIUM",
        "finding": (
            f"Supply fan VFD signal at or above {params.get('fan_max_threshold_pct', 99)}% for "
            f"{pct_at_max:.0f}% of the sample period. Mean speed: {sf.mean():.1f}%. "
            f"Min speed: {sf.min():.0f}%. This may indicate duct static pressure issues, "
            f"oversized duct static setpoint, or zones calling for maximum air."
        ),
        "recommendation": (
            "Review duct static pressure setpoint — verify it matches TAB report values. "
            "Check for duct leakage, disconnected ductwork, or stuck-open VAV dampers. "
            "Evaluate duct static pressure optimization (trim-and-respond) if not already active. "
            "Verify VFD is programmed correctly and not bypassed."
        ),
        "data": {
            "pct_at_max": round(pct_at_max, 1),
            "mean_speed_pct": round(float(sf.mean()), 1),
            "min_speed_pct": round(float(sf.min()), 1),
        },
    }


# ── Energy Waste Quantification ───────────────────────────────────────────────

def quantify_energy_waste(df, cols, params, faults):
    """Compute energy waste estimates for quantifiable faults.

    Returns dict of fault_id -> waste metrics, plus raw Python values
    for the Validation Column in the Calculations sheet.
    """
    results = {}
    interval_hrs = params["interval_min"] / 60
    sample_days = params["sample_days"]
    annual_factor = 365 / sample_days if sample_days > 0 else 1

    # ── Simultaneous heating/cooling waste ────────────────────────────────
    # Methodology: During simultaneous periods, the CHW valve opens to fight
    # the stuck HW valve. The CHW energy is 100% waste (it wouldn't be needed
    # if HW were closed). We estimate cooling load from CHW valve position
    # as a fraction of design capacity: Q_clg = (CHW_pct/100) * design_tons.
    # The heating energy dumped into the airstream (and then removed by CHW)
    # is estimated at the same magnitude converted to therms.
    simul_fault = next((f for f in faults if f["id"] == "SIMUL-HTG-CLG"), None)
    if simul_fault:
        hw = df[cols["hw_valve"]]
        chw = df[cols["chw_valve"]]
        mask = simul_fault.get("simul_mask", (hw > 5) & (chw > 5))

        design_tons = params.get("cooling_tons", 10)
        kw_per_ton = params.get("kw_per_ton", 1.0)
        avg_chw_pct = chw[mask].mean()
        simul_hours = mask.sum() * interval_hrs

        # Cooling waste: valve fraction × design capacity × hours
        sample_tonhrs_clg = (avg_chw_pct / 100) * design_tons * simul_hours
        annual_tonhrs_clg = sample_tonhrs_clg * annual_factor
        annual_kwh_clg = annual_tonhrs_clg * kw_per_ton
        annual_cost_clg = annual_kwh_clg * params["elec_rate"]

        # Heating waste: same ton-hrs converted to BTU then therms
        sample_btu_htg = sample_tonhrs_clg * 12000
        annual_therms_htg = (sample_btu_htg * annual_factor) / 100000
        heating_rate = params.get("heating_rate_per_therm", 1.50)
        annual_cost_htg = annual_therms_htg * heating_rate

        results["SIMUL-HTG-CLG"] = {
            "avg_chw_pct": round(avg_chw_pct, 1),
            "simul_hours": round(simul_hours, 1),
            "sample_tonhrs_clg": round(sample_tonhrs_clg, 1),
            "annual_tonhrs_clg": round(annual_tonhrs_clg, 1),
            "annual_kwh_cooling": round(annual_kwh_clg),
            "annual_cost_cooling": round(annual_cost_clg),
            "annual_therms_heating": round(annual_therms_htg),
            "annual_cost_heating": round(annual_cost_htg),
            "annual_total_cost": round(annual_cost_clg + annual_cost_htg),
        }

    # ── Economizer missed opportunity ─────────────────────────────────────
    econ_fault = next((f for f in faults if f["id"] == "ECON-STUCK"), None)
    if econ_fault:
        oat = df[cols["oat"]]
        rat = df[cols["rat"]]
        fc_mask = econ_fault.get("fc_mask", (oat >= 45) & (oat <= 70))
        fc_df = df[fc_mask]

        if len(fc_df) > 0:
            # Potential cooling displacement: if economizer were at 100%,
            # MAT would approach OAT. Delta vs current MAT at 10%.
            cfm = params.get("design_cfm", 5000)
            # Current MAT estimate at 10% OA
            current_mat = 0.10 * oat[fc_mask] + 0.90 * rat[fc_mask]
            # Ideal MAT at ~80% OA (realistic modulation)
            ideal_mat = 0.80 * oat[fc_mask] + 0.20 * rat[fc_mask]
            delta_t = current_mat - ideal_mat  # degrees cooler supply possible

            saved_btu_per_interval = 1.08 * cfm * delta_t.clip(lower=0) * interval_hrs
            total_saved_btu = saved_btu_per_interval.sum()
            total_saved_tonhrs = total_saved_btu / 12000
            annual_saved_tonhrs = total_saved_tonhrs * annual_factor
            annual_kwh = annual_saved_tonhrs * params.get("kw_per_ton", 1.0)
            annual_cost = annual_kwh * params["elec_rate"]

            results["ECON-STUCK"] = {
                "sample_saved_btu": round(total_saved_btu),
                "annual_saved_tonhrs": round(annual_saved_tonhrs, 1),
                "annual_kwh": round(annual_kwh),
                "annual_cost": round(annual_cost),
            }

    return results


# ── Excel Workbook Generation ─────────────────────────────────────────────────

def build_workbook(df, cols, params, faults, waste, output_path):
    """Generate the analysis Excel workbook."""
    wb = Workbook()
    wb.remove(wb.active)

    ws_summary = wb.create_sheet("Summary", 0)
    build_summary_sheet(ws_summary, params, faults, waste)

    ws_calc = wb.create_sheet("Calculations", 1)
    build_calculations_sheet(ws_calc, df, cols, params, faults, waste)

    ws_charts = wb.create_sheet("Fault Charts", 2)
    build_charts_sheet(ws_charts, df, cols, params, faults)

    ws_data = wb.create_sheet("Trend Data", 3)
    build_data_sheet(ws_data, df, cols, params)

    wb.save(output_path)
    print(f"  Workbook saved: {output_path}")


def build_summary_sheet(ws, params, faults, waste):
    """Summary of all findings, severity, and recommendations."""
    ws.sheet_properties.tabColor = "4472C4"

    # Title block
    ws["A1"] = "AHU TREND DATA ANALYSIS — FAULT DETECTION REPORT"
    ws["A1"].font = Font(size=16, bold=True, color="1F3864")
    ws["A2"] = f"Equipment: {params.get('equipment_name', 'AHU-1')}"
    ws["A2"].font = Font(size=12, bold=True)
    ws["A3"] = f"Project: {params.get('project_name', '')}"
    ws["A4"] = f"Data Period: {params.get('date_start', '')} to {params.get('date_end', '')}"
    ws["A5"] = f"Sample Days: {params['sample_days']:.0f} | Intervals: {params['total_intervals']} | Interval: {params['interval_min']} min"
    ws["A6"] = f"Analysis Date: {pd.Timestamp.now().strftime('%Y-%m-%d')}"

    # Fault summary table
    row = 8
    section_header(ws, row, "FAULTS DETECTED", cols=7)
    row += 1

    headers = ["#", "Fault ID", "Title", "Severity", "Finding", "Recommendation", "Est. Annual Cost"]
    for c, h in enumerate(headers, 1):
        ws.cell(row, c, h)
    style_header_row(ws, row, len(headers))
    row += 1
    start_data = row

    severity_colors = {
        "CRITICAL": PatternFill("solid", fgColor="FF6B6B"),
        "HIGH": PatternFill("solid", fgColor="FFA500"),
        "MEDIUM": PatternFill("solid", fgColor="FFD700"),
        "LOW": PatternFill("solid", fgColor="90EE90"),
    }

    for i, fault in enumerate(faults, 1):
        ws.cell(row, 1, i)
        ws.cell(row, 2, fault["id"])
        ws.cell(row, 3, fault["title"])
        sev_cell = ws.cell(row, 4, fault["severity"])
        sev_cell.fill = severity_colors.get(fault["severity"], PatternFill())
        sev_cell.font = Font(bold=True)
        ws.cell(row, 5, fault["finding"])
        ws.cell(row, 6, fault["recommendation"])
        # Annual cost from waste calculations
        w = waste.get(fault["id"], {})
        cost = w.get("annual_total_cost", w.get("annual_cost", ""))
        if cost:
            c = ws.cell(row, 7, cost)
            c.number_format = dollar_fmt
            c.font = Font(bold=True, color="008000")
        row += 1

    style_range(ws, start_data, row - 1, len(headers))

    # Total savings
    if any(waste.values()):
        row += 1
        ws.cell(row, 6, "TOTAL ESTIMATED ANNUAL SAVINGS:").font = Font(bold=True, size=12)
        total = sum(
            w.get("annual_total_cost", w.get("annual_cost", 0))
            for w in waste.values() if isinstance(w.get("annual_total_cost", w.get("annual_cost", 0)), (int, float))
        )
        c = ws.cell(row, 7, total)
        c.number_format = dollar_fmt
        c.font = Font(bold=True, color="008000", size=14)

    set_col_widths(ws, [5, 22, 38, 12, 70, 60, 18])

    # Wrap text on finding/recommendation columns
    for r in range(start_data, row):
        for col in [5, 6]:
            ws.cell(r, col).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row, 6).alignment = Alignment(horizontal="right")


def build_calculations_sheet(ws, df, cols, params, faults, waste):
    """Auditable calculations with Excel formulas and Python validation."""
    ws.sheet_properties.tabColor = "008000"

    row = 1
    ws.cell(row, 1, "CALCULATIONS — ENGINEERING METHODOLOGY & FORMULAS").font = Font(size=14, bold=True, color="1F3864")
    row += 2

    # ── Color legend ──────────────────────────────────────────────────────
    row = add_color_legend(ws, row)
    row += 1

    # ── Assumptions Section ───────────────────────────────────────────────
    section_header(ws, row, "ASSUMPTIONS (User-Editable Inputs)", cols=6)
    row += 1
    headers = ["Parameter", "Value", "Unit", "Source"]
    for c, h in enumerate(headers, 1):
        ws.cell(row, c, h)
    style_header_row(ws, row, len(headers))
    row += 1
    assume_start = row

    assumptions = [
        ("Design CFM", params.get("design_cfm", 5000), "CFM", "VERIFY — confirm from TAB report or nameplate"),
        ("Electricity Rate", params["elec_rate"], "$/kWh", "VERIFY — confirm from utility bill"),
        ("Heating Rate", params.get("heating_rate_per_therm", 1.50), "$/therm", "VERIFY — confirm from utility bill"),
        ("Chiller Efficiency", params.get("kw_per_ton", 1.0), "kW/ton", "VERIFY — nameplate or performance data"),
        ("Data Interval", params["interval_min"], "minutes", "BAS trend configuration"),
        ("Sample Period", params["sample_days"], "days", f"Data: {params.get('date_start','')} to {params.get('date_end','')}"),
        ("Annualization Factor", None, "—", "365 / Sample Period (calculated)"),
        ("Economizer Free-Cool Low", params.get("econ_lo_f", 45), "°F", "ASHRAE 90.1 typical"),
        ("Economizer Free-Cool High", params.get("econ_hi_f", 70), "°F", "ASHRAE 90.1 typical"),
        ("Economizer % OA (ideal)", 0.80, "fraction", "Conservative (not 100% — accounts for min RA)"),
        ("Economizer % OA (current)", 0.10, "fraction", "From trend data — damper locked at 10%"),
    ]

    for label, value, unit, source in assumptions:
        ws.cell(row, 1, label).font = black_bold
        if value is not None:
            input_cell(ws, row, 2, value)
        else:
            # Formula: annualization
            formula_cell(ws, row, 2, f"=365/B{row - 1}")
        ws.cell(row, 3, unit)
        src_cell = ws.cell(row, 4, source)
        if "VERIFY" in source:
            src_cell.font = warn_font
        else:
            src_cell.font = note_font
        row += 1

    assume_end = row - 1
    style_range(ws, assume_start, assume_end, 4)

    # Row references for formulas
    r_cfm = assume_start
    r_elec = assume_start + 1
    r_htg_rate = assume_start + 2
    r_kw_ton = assume_start + 3
    r_interval = assume_start + 4
    r_sample = assume_start + 5
    r_annual = assume_start + 6
    r_econ_lo = assume_start + 7
    r_econ_hi = assume_start + 8
    r_econ_ideal = assume_start + 9
    r_econ_current = assume_start + 10

    row += 2

    # ── Simultaneous Heating/Cooling Calculation ──────────────────────────
    simul_fault = next((f for f in faults if f["id"] == "SIMUL-HTG-CLG"), None)
    simul_waste = waste.get("SIMUL-HTG-CLG", {})
    if simul_fault and simul_waste:
        section_header(ws, row, "SIMULTANEOUS HEATING/COOLING — ENERGY WASTE", cols=6)
        row += 1

        ws.cell(row, 1, (
            "Methodology: During simultaneous operation, the CHW valve opens solely to "
            "counteract the stuck HW valve. Cooling waste is estimated as "
            "(avg CHW valve % / 100) × design cooling capacity × hours. "
            "Heating waste is the same energy in BTU converted to therms. "
            "This valve-position method is appropriate when DAT is controlled near setpoint "
            "by the fighting valves (DAT excess ≈ 0)."
        )).font = note_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row, 1).alignment = Alignment(wrap_text=True)
        row += 2

        calc_headers = ["Step", "Parameter", "Formula / Value", "Result", "Unit", "Python Check"]
        for c, h in enumerate(calc_headers, 1):
            ws.cell(row, c, h)
        style_header_row(ws, row, len(calc_headers))
        row += 1
        calc_start = row

        # Step 1: Simultaneous intervals
        ws.cell(row, 1, 1)
        ws.cell(row, 2, "Simultaneous intervals (from data)")
        input_cell(ws, row, 3, simul_fault["data"]["intervals_affected"], highlight=True)
        formula_cell(ws, row, 4, f"=C{row}")
        ws.cell(row, 5, "intervals")
        ws.cell(row, 6, simul_fault["data"]["intervals_affected"]).font = note_font
        r_simul_n = row
        row += 1

        # Step 2: Simultaneous hours
        ws.cell(row, 1, 2)
        ws.cell(row, 2, "Simultaneous hours")
        formula_cell(ws, row, 3, f"=C{r_simul_n}*B{r_interval}/60")
        formula_cell(ws, row, 4, f"=C{row}", fmt=dec1_fmt)
        ws.cell(row, 5, "hours")
        ws.cell(row, 6, simul_waste["simul_hours"]).font = note_font
        r_simul_hrs = row
        row += 1

        # Step 3: Avg CHW valve during simultaneous periods
        ws.cell(row, 1, 3)
        ws.cell(row, 2, "Avg CHW valve position during simultaneous periods")
        input_cell(ws, row, 3, simul_waste["avg_chw_pct"], highlight=True)
        formula_cell(ws, row, 4, f"=C{row}")
        ws.cell(row, 5, "%")
        ws.cell(row, 6, simul_waste["avg_chw_pct"]).font = note_font
        r_avg_chw = row
        row += 1

        # Step 4: Design cooling capacity
        ws.cell(row, 1, 4)
        ws.cell(row, 2, "Design cooling capacity")
        input_cell(ws, row, 3, params.get("cooling_tons", 10), highlight=True)
        formula_cell(ws, row, 4, f"=C{row}")
        ws.cell(row, 5, "tons")
        src = ws.cell(row, 6, "VERIFY — confirm from nameplate or schedules")
        src.font = warn_font
        r_design_tons = row
        row += 1

        # Step 5: Sample cooling waste (ton-hrs)
        ws.cell(row, 1, 5)
        ws.cell(row, 2, "Sample period cooling waste")
        formula_cell(ws, row, 3, f"=(C{r_avg_chw}/100)*C{r_design_tons}*D{r_simul_hrs}")
        formula_cell(ws, row, 4, f"=C{row}", fmt=dec1_fmt)
        ws.cell(row, 5, "ton-hrs")
        ws.cell(row, 6, simul_waste["sample_tonhrs_clg"]).font = note_font
        r_sample_tonhrs = row
        row += 1

        # Step 6: Annual cooling waste (ton-hrs)
        ws.cell(row, 1, 6)
        ws.cell(row, 2, "Annual cooling waste")
        formula_cell(ws, row, 3, f"=C{r_sample_tonhrs}*B{r_annual}")
        formula_cell(ws, row, 4, f"=C{row}", fmt=dec1_fmt)
        ws.cell(row, 5, "ton-hrs/yr")
        ws.cell(row, 6, simul_waste["annual_tonhrs_clg"]).font = note_font
        r_annual_tonhrs = row
        row += 1

        # Step 7: Annual cooling kWh
        ws.cell(row, 1, 7)
        ws.cell(row, 2, "Annual cooling electricity waste")
        formula_cell(ws, row, 3, f"=C{r_annual_tonhrs}*B{r_kw_ton}")
        formula_cell(ws, row, 4, f"=C{row}", fmt=kwh_fmt)
        ws.cell(row, 5, "kWh/yr")
        ws.cell(row, 6, simul_waste["annual_kwh_cooling"]).font = note_font
        r_kwh_clg = row
        row += 1

        # Step 8: Annual cooling cost
        ws.cell(row, 1, 8)
        ws.cell(row, 2, "Annual cooling cost waste")
        formula_cell(ws, row, 3, f"=C{r_kwh_clg}*B{r_elec}")
        formula_cell(ws, row, 4, f"=C{row}", fmt=dollar_fmt)
        ws.cell(row, 5, "$/yr")
        ws.cell(row, 6, simul_waste["annual_cost_cooling"]).font = note_font
        r_cost_clg = row
        row += 1

        # Step 9: Annual heating waste (therms)
        ws.cell(row, 1, 9)
        ws.cell(row, 2, "Annual heating waste (= cooling BTU ÷ 100,000)")
        formula_cell(ws, row, 3, f"=C{r_sample_tonhrs}*12000*B{r_annual}/100000")
        formula_cell(ws, row, 4, f"=C{row}", fmt=num_fmt)
        ws.cell(row, 5, "therms/yr")
        ws.cell(row, 6, simul_waste["annual_therms_heating"]).font = note_font
        r_therms = row
        row += 1

        # Step 10: Annual heating cost
        ws.cell(row, 1, 10)
        ws.cell(row, 2, "Annual heating cost waste")
        formula_cell(ws, row, 3, f"=C{r_therms}*B{r_htg_rate}")
        formula_cell(ws, row, 4, f"=C{row}", fmt=dollar_fmt)
        ws.cell(row, 5, "$/yr")
        ws.cell(row, 6, simul_waste["annual_cost_heating"]).font = note_font
        r_cost_htg = row
        row += 1

        # Step 11: Total annual cost
        ws.cell(row, 1, 11)
        ws.cell(row, 2, "TOTAL ANNUAL WASTE — Simultaneous Htg/Clg").font = Font(bold=True, size=11)
        formula_cell(ws, row, 3, f"=C{r_cost_clg}+C{r_cost_htg}")
        c = formula_cell(ws, row, 4, f"=C{row}", fmt=dollar_fmt)
        c.font = Font(bold=True, color="008000", size=12)
        ws.cell(row, 5, "$/yr")
        ws.cell(row, 6, simul_waste["annual_total_cost"]).font = note_font
        row += 1

        style_range(ws, calc_start, row - 1, 6)
        row += 1
        ws.cell(row, 6, "← Python-computed validation values").font = note_font
        row += 2

    # ── Economizer Missed Opportunity Calculation ─────────────────────────
    econ_fault = next((f for f in faults if f["id"] == "ECON-STUCK"), None)
    econ_waste = waste.get("ECON-STUCK", {})
    if econ_fault and econ_waste:
        section_header(ws, row, "ECONOMIZER MISSED FREE-COOLING — ENERGY SAVINGS", cols=6)
        row += 1

        ws.cell(row, 1, (
            "Methodology: Compare mixed air temperature at current 10% OA fraction vs. "
            "ideal 80% OA fraction during free-cooling conditions. The delta represents "
            "cooling that the economizer could have provided for free. "
            "Q = 1.08 × CFM × ΔT (BTU/hr), converted to ton-hrs then kWh."
        )).font = note_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row, 1).alignment = Alignment(wrap_text=True)
        row += 2

        calc_headers = ["Step", "Parameter", "Formula / Value", "Result", "Unit", "Python Check"]
        for c, h in enumerate(calc_headers, 1):
            ws.cell(row, c, h)
        style_header_row(ws, row, len(calc_headers))
        row += 1
        ec_start = row

        # Free cooling hours
        ws.cell(row, 1, 1)
        ws.cell(row, 2, "Free-cooling hours in sample")
        input_cell(ws, row, 3, econ_fault["data"]["free_cooling_hours"], highlight=True)
        formula_cell(ws, row, 4, f"=C{row}")
        ws.cell(row, 5, "hours")
        ws.cell(row, 6, econ_fault["data"]["free_cooling_hours"]).font = note_font
        r_fc_hrs = row
        row += 1

        # Avg delta-T (MAT improvement)
        oat_fc = df[cols["oat"]]
        rat_fc = df[cols["rat"]]
        fc_mask = (oat_fc >= params.get("econ_lo_f", 45)) & (oat_fc <= params.get("econ_hi_f", 70))
        current_mat = 0.10 * oat_fc[fc_mask] + 0.90 * rat_fc[fc_mask]
        ideal_mat = 0.80 * oat_fc[fc_mask] + 0.20 * rat_fc[fc_mask]
        avg_dt = (current_mat - ideal_mat).clip(lower=0).mean()

        ws.cell(row, 1, 2)
        ws.cell(row, 2, "Avg MAT improvement (current vs ideal)")
        input_cell(ws, row, 3, round(avg_dt, 2), highlight=True)
        formula_cell(ws, row, 4, f"=C{row}")
        ws.cell(row, 5, "°F")
        ws.cell(row, 6, round(avg_dt, 2)).font = note_font
        r_econ_dt = row
        row += 1

        # Sample saved BTU
        ws.cell(row, 1, 3)
        ws.cell(row, 2, "Sample period saved BTU")
        formula_cell(ws, row, 3, f"=1.08*B{r_cfm}*C{r_econ_dt}*C{r_fc_hrs}")
        formula_cell(ws, row, 4, f"=C{row}", fmt=num_fmt)
        ws.cell(row, 5, "BTU")
        ws.cell(row, 6, econ_waste["sample_saved_btu"]).font = note_font
        r_econ_btu = row
        row += 1

        # Annual ton-hours saved
        ws.cell(row, 1, 4)
        ws.cell(row, 2, "Annual cooling saved")
        formula_cell(ws, row, 3, f"=C{r_econ_btu}*B{r_annual}/12000")
        formula_cell(ws, row, 4, f"=C{row}", fmt=dec1_fmt)
        ws.cell(row, 5, "ton-hrs/yr")
        ws.cell(row, 6, econ_waste["annual_saved_tonhrs"]).font = note_font
        r_econ_tonhrs = row
        row += 1

        # Annual kWh
        ws.cell(row, 1, 5)
        ws.cell(row, 2, "Annual electricity saved")
        formula_cell(ws, row, 3, f"=C{r_econ_tonhrs}*B{r_kw_ton}")
        formula_cell(ws, row, 4, f"=C{row}", fmt=kwh_fmt)
        ws.cell(row, 5, "kWh/yr")
        ws.cell(row, 6, econ_waste["annual_kwh"]).font = note_font
        r_econ_kwh = row
        row += 1

        # Annual cost
        ws.cell(row, 1, 6)
        ws.cell(row, 2, "TOTAL ANNUAL SAVINGS — Economizer Fix").font = Font(bold=True, size=11)
        formula_cell(ws, row, 3, f"=C{r_econ_kwh}*B{r_elec}")
        c = formula_cell(ws, row, 4, f"=C{row}", fmt=dollar_fmt)
        c.font = Font(bold=True, color="008000", size=12)
        ws.cell(row, 5, "$/yr")
        ws.cell(row, 6, econ_waste["annual_cost"]).font = note_font
        row += 1

        style_range(ws, ec_start, row - 1, 6)
        row += 1
        ws.cell(row, 6, "← Python-computed validation values").font = note_font

    set_col_widths(ws, [8, 45, 30, 18, 14, 16])


def build_charts_sheet(ws, df, cols, params, faults):
    """Visualization charts showing fault signatures."""
    ws.sheet_properties.tabColor = "FFA500"

    # Prepare a subset dataframe with key columns for charting
    chart_cols = {
        "Timestamp": cols["datetime"],
        "OAT": cols["oat"],
        "MAT": cols["mat"],
        "DAT": cols["dat"],
        "DAT SP": cols["dat_sp"],
        "HW Valve": cols["hw_valve"],
        "CHW Valve": cols["chw_valve"],
        "Econ Dmpr": cols["econ_dmpr"],
        "SF Speed": cols["sf_spd"],
        "Humidity": cols["humidity"],
    }

    # Write chart data starting at A1
    for c, (label, col_name) in enumerate(chart_cols.items(), 1):
        ws.cell(1, c, label).font = header_font
        ws.cell(1, c).fill = header_fill
        ws.cell(1, c).alignment = Alignment(horizontal="center")

    n_rows = len(df)
    for r_idx in range(n_rows):
        for c_idx, (label, col_name) in enumerate(chart_cols.items(), 1):
            val = df.iloc[r_idx][col_name] if col_name in df.columns else ""
            ws.cell(r_idx + 2, c_idx, val)

    data_end = n_rows + 1
    n_chart_cols = len(chart_cols)

    # Chart placement column (after data)
    chart_col = get_column_letter(n_chart_cols + 2)

    # ── Chart 1: Temperature profiles ─────────────────────────────────────
    c1 = LineChart()
    c1.title = "AHU-1 Temperature Profiles"
    c1.y_axis.title = "Temperature (°F)"
    c1.x_axis.title = "Time"
    c1.width = 30
    c1.height = 15
    c1.style = 10

    # OAT, MAT, DAT, DAT SP (cols 2,3,4,5)
    for col_idx, color in [(2, "0070C0"), (3, "7030A0"), (4, "FF0000"), (5, "00B050")]:
        data = Reference(ws, min_col=col_idx, min_row=1, max_row=data_end)
        c1.add_data(data, titles_from_data=True)
        c1.series[-1].graphicalProperties.line.width = 15000
        c1.series[-1].graphicalProperties.line.solidFill = color

    cats = Reference(ws, min_col=1, min_row=2, max_row=data_end)
    c1.set_categories(cats)
    c1.x_axis.delete = True  # Too many labels
    ws.add_chart(c1, f"{chart_col}1")

    # ── Chart 2: Valve positions (HW vs CHW) ──────────────────────────────
    c2 = LineChart()
    c2.title = "AHU-1 Valve Positions — Simultaneous Heating/Cooling"
    c2.y_axis.title = "Valve Position (%)"
    c2.x_axis.title = "Time"
    c2.width = 30
    c2.height = 15
    c2.style = 10

    for col_idx, color in [(6, "FF0000"), (7, "0070C0")]:
        data = Reference(ws, min_col=col_idx, min_row=1, max_row=data_end)
        c2.add_data(data, titles_from_data=True)
        c2.series[-1].graphicalProperties.line.width = 15000
        c2.series[-1].graphicalProperties.line.solidFill = color

    c2.set_categories(cats)
    c2.x_axis.delete = True
    ws.add_chart(c2, f"{chart_col}17")

    # ── Chart 3: Economizer + OAT ─────────────────────────────────────────
    c3 = LineChart()
    c3.title = "AHU-1 Economizer Damper vs Outside Air Temp"
    c3.y_axis.title = "OAT (°F) / Damper (%)"
    c3.width = 30
    c3.height = 15
    c3.style = 10

    for col_idx, color in [(2, "0070C0"), (8, "FFA500")]:
        data = Reference(ws, min_col=col_idx, min_row=1, max_row=data_end)
        c3.add_data(data, titles_from_data=True)
        c3.series[-1].graphicalProperties.line.width = 15000
        c3.series[-1].graphicalProperties.line.solidFill = color

    c3.set_categories(cats)
    c3.x_axis.delete = True
    ws.add_chart(c3, f"{chart_col}33")

    # ── Chart 4: Humidity vs setpoint ─────────────────────────────────────
    c4 = LineChart()
    c4.title = "AHU-1 Space Humidity vs Setpoint (40%)"
    c4.y_axis.title = "Humidity (%)"
    c4.width = 30
    c4.height = 15
    c4.style = 10

    data = Reference(ws, min_col=10, min_row=1, max_row=data_end)
    c4.add_data(data, titles_from_data=True)
    c4.series[-1].graphicalProperties.line.width = 15000
    c4.series[-1].graphicalProperties.line.solidFill = "0070C0"

    c4.set_categories(cats)
    c4.x_axis.delete = True
    ws.add_chart(c4, f"{chart_col}49")


def build_data_sheet(ws, df, cols, params):
    """Raw trend data for reference — key columns only."""
    ws.sheet_properties.tabColor = "808080"

    key_cols = [
        cols["datetime"], cols["oat"], cols["mat"], cols["dat"], cols["dat_sp"],
        cols["rat"], cols["hw_valve"], cols["chw_valve"], cols["econ_dmpr"],
        cols["sf_spd"], cols["rf_spd"], cols["humidity"], cols["hum_sp"],
    ]
    key_cols = [c for c in key_cols if c in df.columns]

    for c_idx, col_name in enumerate(key_cols, 1):
        ws.cell(1, c_idx, col_name).font = header_font
        ws.cell(1, c_idx).fill = header_fill
        ws.cell(1, c_idx).alignment = Alignment(horizontal="center", wrap_text=True)

    for r_idx in range(len(df)):
        for c_idx, col_name in enumerate(key_cols, 1):
            ws.cell(r_idx + 2, c_idx, df.iloc[r_idx][col_name])

    # Auto-width
    for c_idx in range(1, len(key_cols) + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 18


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="AHU Trend Data Analysis — Fault Detection & Energy Waste Quantification"
    )
    parser.add_argument("--csv", type=str, help="Path to cleaned CSV trend data")
    parser.add_argument("--config", type=str, help="Path to JSON config file (overrides CLI args)")
    parser.add_argument("--output", type=str, required=True, help="Output Excel file path")

    # Column mappings (override defaults)
    for key, default in DEFAULT_COLS.items():
        parser.add_argument(f"--col-{key.replace('_', '-')}", type=str, default=default)

    # Analysis parameters
    parser.add_argument("--design-cfm", type=float, default=5000)
    parser.add_argument("--elec-rate", type=float, default=0.10)
    parser.add_argument("--heating-rate", type=float, default=1.50, help="$/therm")
    parser.add_argument("--kw-per-ton", type=float, default=1.0)
    parser.add_argument("--cooling-tons", type=float, default=10.0)
    parser.add_argument("--econ-lo", type=float, default=45, help="Economizer low limit °F")
    parser.add_argument("--econ-hi", type=float, default=70, help="Economizer high limit °F")
    parser.add_argument("--equipment-name", type=str, default="AHU-1")
    parser.add_argument("--project-name", type=str, default="")
    parser.add_argument("--museum", action="store_true", help="Flag for museum (artifact preservation)")

    args = parser.parse_args()

    # Load config if provided
    if args.config:
        with open(args.config) as f:
            cfg = json.load(f)
        csv_path = cfg.get("csv", args.csv)
        col_map = cfg.get("columns", {})
        for key in DEFAULT_COLS:
            cli_key = f"col_{key}"
            if key in col_map:
                setattr(args, cli_key, col_map[key])
        for k, v in cfg.get("params", {}).items():
            setattr(args, k.replace("-", "_"), v)
    else:
        csv_path = args.csv

    if not csv_path:
        print("ERROR: --csv or --config with csv path required")
        sys.exit(1)

    # Build column mapping
    cols = {}
    for key in DEFAULT_COLS:
        cols[key] = getattr(args, f"col_{key}", DEFAULT_COLS[key])

    print("=" * 80)
    print("AHU TREND DATA ANALYSIS TOOL")
    print("=" * 80)
    print(f"Reading: {csv_path}")

    df = pd.read_csv(csv_path)
    df[cols["datetime"]] = pd.to_datetime(df[cols["datetime"]])

    date_start = df[cols["datetime"]].min().strftime("%Y-%m-%d")
    date_end = df[cols["datetime"]].max().strftime("%Y-%m-%d")
    sample_days = (df[cols["datetime"]].max() - df[cols["datetime"]].min()).total_seconds() / 86400
    interval_min = 15  # Default; detect from data
    if len(df) > 1:
        diffs = df[cols["datetime"]].diff().dropna().dt.total_seconds() / 60
        interval_min = diffs.median()

    params = {
        "design_cfm": args.design_cfm,
        "elec_rate": args.elec_rate,
        "heating_rate_per_therm": args.heating_rate,
        "kw_per_ton": args.kw_per_ton,
        "cooling_tons": args.cooling_tons,
        "econ_lo_f": args.econ_lo,
        "econ_hi_f": args.econ_hi,
        "interval_min": interval_min,
        "sample_days": sample_days,
        "date_start": date_start,
        "date_end": date_end,
        "total_intervals": len(df),
        "equipment_name": args.equipment_name,
        "project_name": args.project_name,
        "museum": args.museum,
    }

    print(f"Dataset: {len(df)} points, {date_start} to {date_end} ({sample_days:.1f} days)")
    print(f"Interval: {interval_min:.0f} min")
    print()

    # ── Run fault detection ───────────────────────────────────────────────
    faults = []
    detectors = [
        detect_economizer_stuck,
        detect_hw_valve_stuck,
        detect_simul_htg_clg,
        detect_dat_tracking,
        detect_humidification_offline,
        detect_fan_at_max,
    ]

    for detector in detectors:
        try:
            result = detector(df, cols, params)
            if result:
                faults.append(result)
        except Exception as e:
            print(f"  WARNING: {detector.__name__} failed: {e}")

    print(f"FAULTS DETECTED: {len(faults)}")
    for i, f in enumerate(faults, 1):
        print(f"  {i}. [{f['severity']}] {f['id']}: {f['title']}")
    print()

    # ── Quantify energy waste ─────────────────────────────────────────────
    waste = quantify_energy_waste(df, cols, params, faults)
    for fid, w in waste.items():
        cost_key = "annual_total_cost" if "annual_total_cost" in w else "annual_cost"
        print(f"  {fid}: ${w.get(cost_key, 0):,}/yr estimated waste")
    print()

    # ── Generate Excel workbook ───────────────────────────────────────────
    print(f"Generating Excel workbook: {args.output}")
    build_workbook(df, cols, params, faults, waste, args.output)
    print("Done.")


if __name__ == "__main__":
    main()
