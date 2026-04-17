"""Shared Excel styling constants and helpers for ECM calculation tools.

Matches the color/formatting conventions from the Phase 4.1 reference implementation.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# === FONTS ===
blue_font = Font(bold=False, color="0000FF")
blue_bold = Font(bold=True, color="0000FF")
black_bold = Font(bold=True, color="000000")
green_font = Font(color="008000")
header_font = Font(bold=True, color="FFFFFF")
section_font = Font(bold=True, color="1F3864", size=12)
subsection_font = Font(bold=True, color="1F3864", size=11)
note_font = Font(italic=True, color="808080", size=9)
warn_font = Font(bold=True, color="FF0000")

# === FILLS ===
header_fill = PatternFill("solid", fgColor="4472C4")
assumption_fill = PatternFill("solid", fgColor="FFFF00")
section_fill = PatternFill("solid", fgColor="D9E2F3")

# === BORDERS ===
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# === NUMBER FORMATS ===
pct_fmt = '0.0%'
kwh_fmt = '#,##0'
dollar_fmt = '$#,##0'
dollar_fmt2 = '$#,##0.00'
num_fmt = '#,##0'
dec1_fmt = '#,##0.0'
dec2_fmt = '#,##0.00'
dec3_fmt = '#,##0.000'


def style_header_row(ws, row, cols, fill=None, font=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill or header_fill
        cell.font = font or header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border


def style_range(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c).border = thin_border


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def input_cell(ws, row, col, value, fmt=None, highlight=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = blue_font
    if fmt:
        cell.number_format = fmt
    if highlight:
        cell.fill = assumption_fill
    return cell


def formula_cell(ws, row, col, formula, fmt=None):
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = Font(color="000000")
    if fmt:
        cell.number_format = fmt
    return cell


def green_cell(ws, row, col, formula, fmt=None):
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = green_font
    if fmt:
        cell.number_format = fmt
    return cell


def section_header(ws, row, text, cols=8):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = section_font
    cell.fill = section_fill
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).fill = section_fill


def add_color_legend(ws, row):
    ws.cell(row=row, column=1, value="COLOR KEY:").font = warn_font
    ws.cell(row=row + 1, column=1, value="Blue text = hardcoded input assumptions (user-editable)").font = note_font
    ws.cell(row=row + 2, column=1, value="Black text = formulas (auto-calculated)").font = note_font
    ws.cell(row=row + 3, column=1, value="Green text = cross-sheet references").font = note_font
    ws.cell(row=row + 4, column=1, value="Yellow highlight = key assumptions needing verification").font = note_font
    return row + 5
