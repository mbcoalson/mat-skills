"""ECM Economizer Savings Calculator — generates .xlsx with working Excel formulas.

Covers economizer displacement calculations:
  - Monthly economizer hour estimates
  - Per-unit cooling displacement
  - Chiller kW offset

Usage:
  # CLI args (quick, single unit)
  py -3.12 ecm-economizer-savings.py --name "AHU-1" --tons 25 --part-load 0.30 --displacement 0.50 --econ-hours 1950 --elec-rate 0.10 --output econ-savings.xlsx

  # JSON config (reproducible, multi-unit with monthly hours)
  py -3.12 ecm-economizer-savings.py --config inputs.json --output econ-savings.xlsx

JSON config format:
{
  "project": "Example Project",
  "ecm_id": "ECM-01/14",
  "ecm_title": "Building-Wide Economizer Enable",
  "elec_rate": 0.10,
  "kw_per_hp": 0.746,
  "chiller_eer": 10.5,
  "btu_per_ton": 12000,
  "part_load_fraction": 0.30,
  "displacement_fraction": 0.50,
  "monthly_hours": [
    {"month": "January",   "oat_hrs": 400, "useful_hrs": 50,  "note": "Very low cooling load"},
    {"month": "February",  "oat_hrs": 350, "useful_hrs": 50,  "note": "Very low cooling load"},
    {"month": "March",     "oat_hrs": 450, "useful_hrs": 100, "note": "Low cooling load"},
    {"month": "April",     "oat_hrs": 500, "useful_hrs": 250, "note": "Moderate cooling load"},
    {"month": "May",       "oat_hrs": 500, "useful_hrs": 350, "note": "Moderate-high cooling load"},
    {"month": "June",      "oat_hrs": 300, "useful_hrs": 200, "note": "High load but OAT often >70F"},
    {"month": "July",      "oat_hrs": 150, "useful_hrs": 100, "note": "OAT usually >70F"},
    {"month": "August",    "oat_hrs": 200, "useful_hrs": 150, "note": "OAT often >70F"},
    {"month": "September", "oat_hrs": 450, "useful_hrs": 300, "note": "Moderate-high cooling load"},
    {"month": "October",   "oat_hrs": 500, "useful_hrs": 250, "note": "Moderate cooling load"},
    {"month": "November",  "oat_hrs": 400, "useful_hrs": 100, "note": "Low cooling load"},
    {"month": "December",  "oat_hrs": 350, "useful_hrs": 50,  "note": "Very low cooling load"}
  ],
  "equipment": [
    {"name": "AHU-1", "cooling_tons": 25, "note": "CHW coil"},
    {"name": "RTU-3", "cooling_tons": 12, "note": "DX cooling"}
  ],
  "notes": ["ROUGH ESTIMATES. A proper bin analysis with Denver TMY3 data would improve accuracy."],
  "verifications": ["Denver TMY3 bin analysis for economizer hours"]
}
"""
import argparse, json, sys
from pathlib import Path
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent))
from _styles import *


def build_workbook(cfg):
    wb = Workbook()
    equipment = cfg["equipment"]
    elec_rate = cfg.get("elec_rate", 0.10)
    chiller_eer = cfg.get("chiller_eer", 10.5)
    btu_per_ton = cfg.get("btu_per_ton", 12000)
    part_load = cfg.get("part_load_fraction", 0.30)
    displacement = cfg.get("displacement_fraction", 0.50)
    monthly = cfg.get("monthly_hours", None)
    total_econ_hrs = cfg.get("total_econ_hours", 1950)
    ecm_id = cfg.get("ecm_id", "ECM")
    ecm_title = cfg.get("ecm_title", "Economizer Savings")
    project = cfg.get("project", "")

    # ── ASSUMPTIONS SHEET ──
    ws = wb.active
    ws.title = "Assumptions"
    set_col_widths(ws, [38, 18, 14, 50])
    section_header(ws, 1, "Assumptions & Constants", 4)
    for i, h in enumerate(["Parameter", "Value", "Units", "Basis / Source"], 1):
        ws.cell(row=2, column=i, value=h)
    style_header_row(ws, 2, 4)

    assumptions = [
        (3, "Electricity rate", elec_rate, "$/kWh", cfg.get("elec_rate_source", "Verify against utility bills"), dollar_fmt2),
        (4, "Chiller EER", chiller_eer, "", cfg.get("eer_source", "Chiller nameplate or rated efficiency"), dec1_fmt),
        (5, "Btu per ton-hr", btu_per_ton, "Btu/ton-hr", "Standard conversion", num_fmt),
        (6, "Part-load fraction during econ hours", part_load, "", f"{part_load*100:.0f}% of rated capacity (assumption)", pct_fmt),
        (7, "Economizer displacement fraction", displacement, "", f"{displacement*100:.0f}% of cooling load displaced by free cooling", pct_fmt),
    ]
    for row, param, val, units, basis, fmt in assumptions:
        ws.cell(row=row, column=1, value=param)
        input_cell(ws, row, 2, val, fmt)
        ws.cell(row=row, column=3, value=units)
        ws.cell(row=row, column=4, value=basis).font = note_font
    style_range(ws, 3, 7, 4)
    add_color_legend(ws, 9)

    # ── MONTHLY HOURS SHEET (if provided) ──
    monthly_total_cell = None
    if monthly:
        ws_m = wb.create_sheet("Monthly Hours")
        set_col_widths(ws_m, [14, 22, 22, 40])
        section_header(ws_m, 1, "Annual Economizer Hours (Estimated)", 4)

        for note in cfg.get("notes", []):
            ws_m.cell(row=2, column=1, value=note).font = warn_font

        for i, h in enumerate(["Month", "Approx Hrs OAT Range", "Useful Economizer Hrs", "Notes"], 1):
            ws_m.cell(row=3, column=i, value=h)
        style_header_row(ws_m, 3, 4)

        for m_idx, m in enumerate(monthly):
            r = 4 + m_idx
            ws_m.cell(row=r, column=1, value=m["month"])
            input_cell(ws_m, r, 2, m["oat_hrs"], num_fmt)
            input_cell(ws_m, r, 3, m["useful_hrs"], num_fmt)
            ws_m.cell(row=r, column=4, value=m.get("note", "")).font = note_font
        style_range(ws_m, 4, 4 + len(monthly) - 1, 4)

        total_row = 4 + len(monthly)
        ws_m.cell(row=total_row, column=1, value="TOTAL").font = black_bold
        formula_cell(ws_m, total_row, 2, f"=SUM(B4:B{total_row - 1})", num_fmt)
        ws_m.cell(row=total_row, column=2).font = Font(bold=True)
        formula_cell(ws_m, total_row, 3, f"=SUM(C4:C{total_row - 1})", num_fmt)
        ws_m.cell(row=total_row, column=3).font = Font(bold=True)
        for c in range(1, 5):
            ws_m.cell(row=total_row, column=c).border = thin_border
        monthly_total_cell = f"'Monthly Hours'!C{total_row}"
    else:
        # No monthly breakdown — put total hours on Assumptions sheet
        r = 8
        ws.cell(row=r, column=1, value="Total useful economizer hours")
        input_cell(ws, r, 2, total_econ_hrs, num_fmt)
        ws.cell(row=r, column=3, value="hr/yr")
        ws.cell(row=r, column=4, value="Estimate — verify with TMY bin analysis").font = note_font
        style_range(ws, r, r, 4)
        monthly_total_cell = "Assumptions!B8"

    # ── CALCULATIONS SHEET ──
    ws_c = wb.create_sheet("Calculations")
    n_eq = len(equipment)
    set_col_widths(ws_c, [35] + [18] * n_eq + [14])
    section_header(ws_c, 1, f"{ecm_id}: {ecm_title}", 2 + n_eq)

    # Equipment capacities
    r = 3
    ws_c.cell(row=r, column=1, value="Equipment Cooling Capacities").font = subsection_font
    r = 4
    headers = ["Parameter"] + [eq["name"] for eq in equipment] + ["Units"]
    for i, h in enumerate(headers, 1):
        ws_c.cell(row=r, column=i, value=h)
    style_header_row(ws_c, r, len(headers))

    r = 5
    ws_c.cell(row=r, column=1, value="Rated cooling capacity")
    for e_idx, eq in enumerate(equipment):
        input_cell(ws_c, r, 2 + e_idx, eq["cooling_tons"], num_fmt)
    ws_c.cell(row=r, column=2 + n_eq, value="tons")
    style_range(ws_c, 5, 5, len(headers))

    # Calculation steps
    r = 7
    ws_c.cell(row=r, column=1, value="Savings Calculations").font = subsection_font
    r = 8
    calc_headers = ["Calculation Step"] + [eq["name"] for eq in equipment] + ["Units"]
    for i, h in enumerate(calc_headers, 1):
        ws_c.cell(row=r, column=i, value=h)
    style_header_row(ws_c, r, len(calc_headers))

    # Row 9: Avg cooling load = capacity * part-load
    r = 9
    ws_c.cell(row=r, column=1, value="Avg cooling load (capacity x part-load)")
    for e in range(n_eq):
        col = get_column_letter(2 + e)
        formula_cell(ws_c, r, 2 + e, f"={col}5*Assumptions!B6", dec1_fmt)
    ws_c.cell(row=r, column=2 + n_eq, value="tons")

    # Row 10: Cooling displaced = avg load * displacement fraction
    r = 10
    ws_c.cell(row=r, column=1, value="Cooling displaced by economizer")
    for e in range(n_eq):
        col = get_column_letter(2 + e)
        formula_cell(ws_c, r, 2 + e, f"={col}9*Assumptions!B7", dec2_fmt)
    ws_c.cell(row=r, column=2 + n_eq, value="tons")

    # Row 11: kW displaced = tons * Btu/ton / (EER * 1000)
    r = 11
    ws_c.cell(row=r, column=1, value="kW displaced (tons x 12000 / (EER x 1000))")
    for e in range(n_eq):
        col = get_column_letter(2 + e)
        formula_cell(ws_c, r, 2 + e, f"={col}10*Assumptions!B5/(Assumptions!B4*1000)", dec2_fmt)
    ws_c.cell(row=r, column=2 + n_eq, value="kW")

    # Row 12: Annual kWh = kW * useful econ hours
    r = 12
    ws_c.cell(row=r, column=1, value="Annual kWh saved")
    for e in range(n_eq):
        col = get_column_letter(2 + e)
        formula_cell(ws_c, r, 2 + e, f"={col}11*{monthly_total_cell}", kwh_fmt)
    ws_c.cell(row=r, column=2 + n_eq, value="kWh/yr")

    # Row 13: Annual $ = kWh * elec rate
    r = 13
    ws_c.cell(row=r, column=1, value="Annual $ saved")
    for e in range(n_eq):
        col = get_column_letter(2 + e)
        formula_cell(ws_c, r, 2 + e, f"={col}12*Assumptions!B3", dollar_fmt)
    ws_c.cell(row=r, column=2 + n_eq, value="$/yr")
    style_range(ws_c, 9, 13, len(calc_headers))

    # Summary table
    r = 15
    ws_c.cell(row=r, column=1, value=f"{ecm_id} Summary").font = subsection_font
    r = 16
    for i, h in enumerate(["Unit", "kWh/yr Saved", "$/yr Saved"], 1):
        ws_c.cell(row=r, column=i, value=h)
    style_header_row(ws_c, r, 3)

    for e_idx, eq in enumerate(equipment):
        r = 17 + e_idx
        ws_c.cell(row=r, column=1, value=eq["name"])
        col = get_column_letter(2 + e_idx)
        formula_cell(ws_c, r, 2, f"={col}12", kwh_fmt)
        formula_cell(ws_c, r, 3, f"={col}13", dollar_fmt)
    style_range(ws_c, 17, 17 + n_eq - 1, 3)

    # Total
    r = 17 + n_eq
    ws_c.cell(row=r, column=1, value="TOTAL").font = black_bold
    formula_cell(ws_c, r, 2, f"=SUM(B17:B{r - 1})", kwh_fmt)
    ws_c.cell(row=r, column=2).font = Font(bold=True)
    formula_cell(ws_c, r, 3, f"=SUM(C17:C{r - 1})", dollar_fmt)
    ws_c.cell(row=r, column=3).font = Font(bold=True)
    style_range(ws_c, r, r, 3)

    # Confidence caveat
    r = 17 + n_eq + 2
    ws_c.cell(row=r, column=1, value="CONFIDENCE CAVEAT: +/-50% uncertainty. Rough economizer hour estimates and assumed part-load.").font = warn_font

    # Verifications
    verifications = cfg.get("verifications", [])
    if verifications:
        r += 2
        ws_c.cell(row=r, column=1, value="Verification Items").font = subsection_font
        for v_idx, v in enumerate(verifications):
            ws_c.cell(row=r + 1 + v_idx, column=1, value=f"{v_idx + 1}. {v}").font = warn_font

    return wb


def build_config_from_cli(args):
    return {
        "ecm_id": args.ecm_id or "ECM",
        "ecm_title": args.title or "Economizer Savings",
        "project": args.project or "",
        "elec_rate": args.elec_rate,
        "chiller_eer": args.eer,
        "btu_per_ton": 12000,
        "part_load_fraction": args.part_load,
        "displacement_fraction": args.displacement,
        "total_econ_hours": args.econ_hours,
        "equipment": [{"name": args.name, "cooling_tons": args.tons}],
        "verifications": []
    }


def main():
    parser = argparse.ArgumentParser(description="ECM Economizer Savings Calculator — generates .xlsx with Excel formulas")
    parser.add_argument("--config", type=str, help="JSON config file path")
    parser.add_argument("--output", "-o", type=str, default="ecm-economizer-savings.xlsx")
    parser.add_argument("--name", type=str, default="AHU-1")
    parser.add_argument("--tons", type=float, default=25)
    parser.add_argument("--part-load", type=float, default=0.30)
    parser.add_argument("--displacement", type=float, default=0.50)
    parser.add_argument("--econ-hours", type=float, default=1950)
    parser.add_argument("--elec-rate", type=float, default=0.10)
    parser.add_argument("--eer", type=float, default=10.5)
    parser.add_argument("--ecm-id", type=str, default=None)
    parser.add_argument("--title", type=str, default=None)
    parser.add_argument("--project", type=str, default=None)
    args = parser.parse_args()

    if args.config:
        with open(args.config) as f:
            cfg = json.load(f)
    else:
        cfg = build_config_from_cli(args)

    wb = build_workbook(cfg)
    wb.save(args.output)
    print(f"Saved: {args.output}")


if __name__ == "__main__":
    main()
